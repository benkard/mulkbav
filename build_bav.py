# -*- coding: utf-8 -*-
"""
bAV (DYNO) vs. Altersvorsorgedepot vs. privates ETF-Depot
Vollstaendig formelbasiertes Excel-Modell, Rechtsstand 2026.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.chart import BarChart, Reference

import os
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "bAV-DYNO_vs_Altersvorsorgedepot_vs_ETF.xlsx")

FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")
BLACK = Font(name=FONT, size=10)
GREEN = Font(name=FONT, size=10, color="008000")
BOLD = Font(name=FONT, size=10, bold=True)
H1 = Font(name=FONT, size=14, bold=True, color="1F3864")
H2 = Font(name=FONT, size=11, bold=True, color="1F3864")
YELLOW = PatternFill("solid", fgColor="FFFF00")
GREYFILL = PatternFill("solid", fgColor="D9E1F2")
HEADFILL = PatternFill("solid", fgColor="1F3864")
HEADFONT = Font(name=FONT, size=9, bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EUR = '#,##0;(#,##0);"-"'
EUR2 = '#,##0.00;(#,##0.00);"-"'
PCT = '0.00%'
PCT1 = '0.0%'
NUM4 = '0.0000'

wb = openpyxl.Workbook()

# =====================================================================
# 1) EINGABEN
# =====================================================================
ws = wb.active
ws.title = "Eingaben"
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 52
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 78

ws['B1'] = "Eingaben - bAV (DYNO) vs. Altersvorsorgedepot vs. privates ETF-Depot"
ws['B1'].font = H1
ws['B2'] = "Nur die GELB hinterlegten Zellen (blaue Schrift) aendern. Alles andere ist Formel."
ws['B2'].font = Font(name=FONT, size=10, italic=True)

inp = [
    ("A", "PERSON UND ZEITACHSE", None, None, None),
    ("alter",      "Alter heute (vollendete Jahre)",                        40,     "Jahre", "Startpunkt der Ansparphase; Kalenderjahr 0 = 2026."),
    ("rentbeg",    "Auszahlungsbeginn (Alter)",                             67,     "Jahre", "Beginn Auszahlungsphase aller drei Optionen. Altersvorsorgedepot: fruehestens 65 (Reform 2026)."),
    ("planende",   "Ende Auszahlungsplan / Vergleichshorizont (Alter)",     85,     "Jahre", "Auszahlungsplaene laufen bis hierhin. Lebenslange Renten werden ueber denselben Horizont verglichen (Memo-Zeile fuer den Rest)."),
    ("ausscheid",  "Ausscheiden aus dem Job / Vorruhestand (Alter)",        63,     "Jahre", "Ab hier kein Gehalt, keine Beitraege mehr. Kapital verzinst sich weiter bis zum Auszahlungsbeginn."),
    ("tzalter",    "Teilzeit ab Alter",                                     55,     "Jahre", "Auf 999 setzen, wenn keine Teilzeit geplant ist."),
    ("tzfaktor",   "Teilzeitfaktor (Anteil des Vollzeitgehalts)",           0.60,   "",      "0,60 = 60 %-Stelle."),
    ("A", "EINKOMMEN UND STEUER", None, None, None),
    ("gehalt",     "Jahresbruttogehalt heute",                              75000,  "EUR",   "Ohne Einmalzahlungen; ggf. inkl. anteiligem 13. Gehalt."),
    ("gsteig",     "Gehalts- und Lohnsteigerung p.a.",                      0.025,  "",      "Wird auch zur Fortschreibung von BBG, Bezugsgroesse, Durchschnittsentgelt und Rentenwert benutzt."),
    ("stkl",       "Veranlagung: 1 = einzeln (Stkl. 1), 3 = zusammen (Stkl. 3)", 1, "1/3",   "Massgeblich ist die VERANLAGUNGSART, nicht die Lohnsteuerklasse: die Klasse steuert nur den unterjaehrigen Abzug, der Jahresbescheid rechnet ohnehin mit Splitting. 3 = Splittingtarif nach 32a Abs. 5 EStG; dann unbedingt das zvE des Partners eintragen."),
    ("partner",    "zu versteuerndes Einkommen des Partners (nur Stkl. 3)", 0,      "EUR",   "Bei Stkl. 3 wird der Splittingtarif nach 32a Abs. 5 EStG gerechnet: 2 x ESt((zvE + zvE_Partner)/2)."),
    ("kist",       "Kirchensteuersatz (0 / 0,08 / 0,09)",                   0.0,    "",      "8 % in BY und BW, sonst 9 %."),
    ("tarifanp",   "Anpassung der Tarifeckwerte p.a. (kalte Progression)",  0.02,   "",      "Der ganze Tarif wird mit diesem Faktor gestreckt. 0 % = volle kalte Progression."),
    ("kinder",     "Anzahl Kinder (fuer Kinderzulage)",                     0,      "",      "Kinderzulage im Altersvorsorgedepot: 1 EUR je eingezahltem EUR bis 300 EUR je Kind."),
    ("kind_jahre", "Jahre, in denen noch Kinderzulage zusteht",             18,     "Jahre", "Die Kinderzulage ist an den Kindergeldanspruch gekoppelt ( 85 EStG), laeuft also nicht ueber die gesamte Ansparphase. Auf 0 setzen, wenn kein Anspruch mehr besteht."),
    ("kinderlos",  "Kinderlos (Zuschlag Pflegeversicherung)? 1 = ja",       0,      "1/0",   "0,6 Prozentpunkte Zuschlag, allein vom Arbeitnehmer zu tragen ( 55 Abs. 3 SGB XI)."),
    ("A", "BETRIEBLICHE ALTERSVORSORGE (DYNO)", None, None, None),
    ("modus",      "Vorgabe: 1 = Bruttobeitrag, 2 = Nettoaufwand",          1,      "1/2",   "Bei 2 wird der Bruttobeitrag so bestimmt, dass der gewuenschte Nettoaufwand herauskommt (Bisektion im Blatt 'Umrechnung'). Der Nettoaufwand ist der tatsaechlich aufgegebene Konsum und damit die oekonomisch richtige Entscheidungsvariable."),
    ("beitrag",    "Monatsbeitrag bAV brutto (nur bei Modus 1)",            300,    "EUR",   "Bruttoentgeltumwandlung nach 1a BetrAVG."),
    ("netto_ziel", "Monatlicher Nettoaufwand (nur bei Modus 2)",            165,    "EUR",   "Der Betrag, um den Ihr verfuegbares Einkommen im ersten Jahr sinkt. Danach wird der Bruttobeitrag nach der Dynamisierungsregel fortgeschrieben."),
    ("dyn",        "Beitrag mit dem Gehalt dynamisieren? 1 = ja",           1,      "1/0",   "1 = Beitrag steigt mit der Gehaltssteigerung, 0 = nominal konstant."),
    ("agzus",      "Arbeitgeberzuschuss (Anteil des Beitrags)",             0.15,   "",      "Gesetzliches Minimum 15 % ( 1a Abs. 1a BetrAVG). DYNO-Arbeitgeber zahlen oft mehr - hier eintragen, was im Vertrag steht."),
    ("agz_fix",    "Fester AG-Zuschuss zusaetzlich (EUR je Monat)",         0,      "EUR",   "Betragsunabhaengiger Festzuschuss aus Tarifvertrag oder Betriebsvereinbarung - bei der Deutschen Bahn und in vielen Tarifwerken ueblich, dort ZUSAETZLICH zu den 15 % nach  1a Abs. 1a BetrAVG. Er ist arbeitgeberfinanzierte Zuwendung nach  3 Nr. 63 EStG und belegt den 8-%-Topf sowie die 4-%-Beitragsfreiheit ( 1 Abs. 1 S. 1 Nr. 9 SvEV) VORRANGIG; die eigene Entgeltumwandlung wird insoweit verdraengt und teurer."),
    ("agz_fix_dyn","Festen Zuschuss dynamisieren? 1 = ja",                  0,      "1/0",   "0 = nominal konstant (Voreinstellung: Festbetraege stehen meist nominal im Tarifvertrag), 1 = steigt mit der Gehalts- bzw. Tarifsteigerung."),
    ("agz_fix_min","Mindest-Eigenbeitrag fuer den festen Zuschuss (EUR/Monat)", 0,  "EUR",   "Matching-Modelle knuepfen den Festbetrag an eine Mindest-Eigenleistung. 0 = bedingungslos; dann fliesst er in jedem aktiven Jahr, auch bei Beitrag null. Vorsicht: ein wirklich bedingungsloser Zuschuss ist kein Differenzeffekt der Umwandlungsentscheidung und verzerrt den Vergleich zugunsten der bAV."),
    ("agzus_sv",   "Zuschussmodus: 0 = voller Beitrag, 1 = sv-freier Anteil, 2 = spitz", 2, "0/1/2", "Gilt nur fuer den prozentualen Zuschuss; der feste Zuschuss oben ist davon unberuehrt. 2 ist der Gesetzeswortlaut: 1a Abs. 1a BetrAVG schuldet den Zuschuss nur, SOWEIT der Arbeitgeber tatsaechlich Sozialversicherungsbeitraege spart. Liegt Ihr Entgelt ueber beiden Beitragsbemessungsgrenzen, spart er nichts - und schuldet nichts. 1 ist die verbreitete pauschale Praxis, 0 eine freiwillige oder tarifvertragliche Zusage."),
    ("auszform",   "bAV-Auszahlung: 1 = Kapital, 2 = lebenslange Rente",    1,      "1/2",   "Kapital: voll steuerpflichtig im Zuflussjahr, KV/PV auf 120 Monate verteilt ( 229 Abs. 1 S. 3 SGB V)."),
    ("rentfak",    "Rentenfaktor bAV (EUR Monatsrente je 10.000 EUR)",      26,     "EUR",   "Nur relevant bei Auszahlungsform 2. 26 EUR entspricht ca. 3,1 % Entnahmequote."),
    ("rentdyn",    "Dynamik der bAV-Rente p.a.",                            0.01,   "",      "Ueberschussbeteiligung / Fondsentwicklung in der Rentenphase."),
    ("fuenftel",   "Fuenftelregelung auf bAV-Kapital pruefen? 1 = ja",      0,      "1/0",   "34 EStG ist bei planmaessiger Kapitalabfindung aus 3 Nr. 63-Vertraegen nach BFH X R 23/15 nur ausnahmsweise anwendbar. Konservativ: 0."),
    ("A", "KAPITALMARKT UND KOSTEN", None, None, None),
    ("rendite",    "Bruttorendite des ETF p.a. (vor Kosten, vor Steuern)",  0.07,   "",      "Identischer ETF in allen drei Optionen - so war die Frage gestellt."),
    ("k_bav",      "Kosten bAV-Produkt p.a. (Versicherungsmantel + TER)",   0.006,  "",      "DYNO wirbt mit provisionsfrei; realistisch bleiben Verwaltungs- und Mantelkosten. Effektivkosten aus dem Produktinformationsblatt eintragen."),
    ("k_av",       "Kosten Altersvorsorgedepot p.a.",                       0.005,  "",      "Standarddepot: Effektivkosten gesetzlich auf 1,0 % p.a. gedeckelt."),
    ("k_priv",     "Kosten privates ETF-Depot p.a.",                        0.002,  "",      "Reine TER eines breiten Welt-ETF."),
    ("infl",       "Inflation p.a. (fuer die reale Darstellung)",           0.02,   "",      "Diskontierungssatz fuer die Barwerte in Kaufkraft von heute."),
    ("A", "RUHESTAND", None, None, None),
    ("kv_ruhe",    "Krankenversicherung im Ruhestand: 1 = GKV, 2 = PKV",    1,      "1/2",   "Entscheidender Hebel: in der GKV sind Versorgungsbezuege mit dem vollen Beitragssatz belegt, private Altersvorsorge nicht."),
    ("sonst_zve",  "sonstiges zu versteuerndes Einkommen im Ruhestand",     22000,  "EUR",   "Vor allem der steuerpflichtige Teil der gesetzlichen Rente, ohne die hier verglichenen Produkte. Bestimmt den Grenzsteuersatz in der Auszahlungsphase."),
    ("rente_br",   "gesetzliche Bruttorente p.a. (heutige Werte)",          24000,  "EUR",   "Nur fuer den Beitragsbemessungsgrenzen-Deckel in der KV: Rente und Versorgungsbezuege teilen sich EINE Obergrenze ( 223 Abs. 3 SGB V)."),
    ("teilkap",    "Altersvorsorgedepot: Teilkapital zu Beginn",            0.0,    "",      "Bis 30 % sind zu Beginn der Auszahlungsphase als Einmalbetrag moeglich. Gilt in beiden Auszahlungsformen."),
    ("av_auszform","Altersvorsorgedepot: 1 = Auszahlungsplan, 2 = lebenslange Rente", 1, "1/2", "Unabhaengig von der bAV-Auszahlungsform: wer die bAV als Kapital nimmt, kann das Depot trotzdem verrenten. Plan: Entnahme bleibt investiert, Besteuerung der nicht gefoerderten Schicht mit dem halben Unterschiedsbetrag. Rente: Kapital geht an den Versicherer, nicht gefoerderte Schicht nur mit dem Ertragsanteil ( 22 Nr. 5 S. 2 Buchst. a EStG)."),
    ("rentfak_av", "Rentenfaktor Altersvorsorgedepot (EUR je 10.000 EUR)",   26,     "EUR",   "Nur bei Auszahlungsform 2. Voreinstellung wie bei der bAV; ein Depotvertrag kann guenstiger sein als ein Versicherungsmantel - Wert aus dem Angebot eintragen."),
    ("rentdyn_av", "Dynamik der Altersvorsorgedepot-Rente p.a.",             0.01,   "",      "Ueberschussbeteiligung bzw. Fondsentwicklung in der Rentenphase des Depotvertrags."),
]

r = 4
IN = {}
for key, label, val, unit, note in inp:
    if key == "A":
        ws.cell(r, 2, label).font = H2
        ws.cell(r, 2).fill = GREYFILL
        ws.cell(r, 3).fill = GREYFILL
        ws.cell(r, 4).fill = GREYFILL
        r += 1
        continue
    ws.cell(r, 2, label).font = BLACK
    c = ws.cell(r, 3, val)
    c.font = BLUE
    c.fill = YELLOW
    c.border = BOX
    if isinstance(val, float) and val < 1.5 and key not in ("tzfaktor",):
        c.number_format = PCT
    elif key == "tzfaktor":
        c.number_format = PCT1
    elif isinstance(val, int) and abs(val) >= 1000:
        c.number_format = EUR
    ws.cell(r, 4, unit).font = Font(name=FONT, size=9, color="808080")
    ws.cell(r, 5, note).font = Font(name=FONT, size=9, color="595959")
    IN[key] = f"Eingaben!$C${r}"
    r += 1

r += 1
ws.cell(r, 2, "ERGEBNIS AUF EINEN BLICK (Details im Blatt 'Vergleich')").font = H2
ws.cell(r, 2).fill = GREYFILL
ws.cell(r, 3).fill = GREYFILL
res_row = r + 1
OPTS = ["bAV mit DYNO", "Altersvorsorgedepot", "privates ETF-Depot"]
blocks = [
    ("Barwert der Netto-Auszahlungen",
     ["=Vergleich!$B$13", "=Vergleich!$C$13", "=Vergleich!$D$13"], EUR,
     "Summe aller Leistungen nach Steuern und Sozialabgaben, auf heute abgezinst - noch ohne Abzug des eigenen Einsatzes."),
    ("Barwert des Netto-Ertrags",
     ["=Vergleich!$B$15", "=Vergleich!$C$15", "=Vergleich!$D$15"], EUR,
     "Dasselbe abzueglich des Barwerts des Nettoaufwands: was die Option ueber den eigenen Einsatz hinaus abwirft."),
    ("Aequivalente Monatsleistung",
     ["=Vergleich!$B$47", "=Vergleich!$C$47", "=Vergleich!$D$47"], EUR2,
     "Konstante Monatszahlung ueber die Auszahlungsphase mit demselben Wert - vergleichbar mit Ihrem heutigen Nettoeinkommen."),
]
rr_ = res_row
for title, fmls, fmt, note in blocks:
    ws.cell(rr_, 2, f"{title} (in heutiger Kaufkraft)").font = H2
    ws.cell(rr_, 5, note).font = Font(name=FONT, size=9, color="595959")
    rr_ += 1
    for i, lab in enumerate(OPTS):
        ws.cell(rr_, 2, f"     {lab}").font = BLACK
        cc = ws.cell(rr_, 3, fmls[i])
        cc.font = Font(name=FONT, size=11, bold=True, color="008000")
        cc.number_format = fmt
        cc.border = BOX
        rr_ += 1
    rr_ += 1
res_end = rr_

# =====================================================================
# 2) PARAMETER
# =====================================================================
pw = wb.create_sheet("Parameter")
pw.sheet_view.showGridLines = False
pw.column_dimensions['A'].width = 3
pw.column_dimensions['B'].width = 56
pw.column_dimensions['C'].width = 14
pw.column_dimensions['D'].width = 92
pw['B1'] = "Rechengroessen und Tarifkonstanten - Rechtsstand 2026"
pw['B1'].font = H1

par = [
    ("A", "SOZIALVERSICHERUNG 2026", None, None),
    ("bbg_rv",  "Beitragsbemessungsgrenze RV/AV (jaehrlich)",          101400, "SVBezGrV 2026 (BGBl. 2025 I Nr. 116); bundeseinheitlich."),
    ("bbg_kv",  "Beitragsbemessungsgrenze KV/PV (jaehrlich)",           69750, "SVBezGrV 2026; 5.812,50 EUR monatlich."),
    ("bezug",   "Bezugsgroesse (monatlich)",                             3955, "SVBezGrV 2026; 47.460 EUR jaehrlich."),
    ("frei_vb", "Freibetrag Versorgungsbezuege KV (monatlich)",         None,  "1/20 der Bezugsgroesse, 226 Abs. 2 S. 2 SGB V = 197,75 EUR/Monat in 2026. In der PV gilt statt des Freibetrags eine Freigrenze."),
    ("de_rv",   "Vorlaeufiges Durchschnittsentgelt RV",                 51944, "Anlage 1 zu SGB VI, Wert 2026 - Nenner fuer Entgeltpunkte."),
    ("rw",      "Aktueller Rentenwert (ab 1.7.2026)",                   42.52, "Rentenwertbestimmungsverordnung 2026; +4,24 % ggue. 40,79 EUR."),
    ("rv_an",   "Rentenversicherung, Arbeitnehmeranteil",               0.093, "18,6 % haelftig."),
    ("av_an",   "Arbeitslosenversicherung, Arbeitnehmeranteil",         0.013, "2,6 % haelftig."),
    ("kv_allg", "Krankenversicherung, allgemeiner Beitragssatz",        0.146, "241 SGB V."),
    ("kv_zus",  "Durchschnittlicher Zusatzbeitrag 2026",                0.029, "Bekanntmachung des BMG; kassenindividuell abweichend."),
    ("pv_ges",  "Pflegeversicherung, Gesamtbeitragssatz",               0.036, "55 SGB XI. Kinderabschlaege ab dem 2. Kind hier nicht modelliert."),
    ("pv_an",   "Pflegeversicherung, Arbeitnehmeranteil",               0.018, "In Sachsen abweichend (2,3 %)."),
    ("pv_kl",   "Zuschlag fuer Kinderlose (allein AN)",                 0.006, "55 Abs. 3 SGB XI."),
    ("A", "EINKOMMENSTEUERTARIF 2026 ( 32a EStG)", None, None),
    ("gfb",   "Grundfreibetrag / Ende Zone 1",                          12348, "32a Abs. 1 EStG i.d.F. des Steuerfortentwicklungsgesetzes."),
    ("e2",    "Ende Zone 2 (Progressionszone I)",                       17799, "Grenzsteuersatz steigt von 14 % auf 23,97 %."),
    ("e3",    "Ende Zone 3 (Progressionszone II)",                      69878, "Grenzsteuersatz steigt von 23,97 % auf 42 %."),
    ("e4",    "Beginn Zone 5 (Reichensteuer)",                         277825, "Ab 277.826 EUR gilt 45 %."),
    ("a2",    "Koeffizient a2 (Zone 2)",                               914.51, "ESt = (a2*y + b2)*y mit y = (zvE - Grundfreibetrag)/10.000."),
    ("b2",    "Koeffizient b2 (Zone 2)",                                 1400, ""),
    ("a3",    "Koeffizient a3 (Zone 3)",                                173.10, "ESt = (a3*z + b3)*z + c3 mit z = (zvE - 17.799)/10.000."),
    ("b3",    "Koeffizient b3 (Zone 3)",                                  2397, ""),
    ("c3",    "Konstante c3 (Zone 3)",                                 1034.87, ""),
    ("c4",    "Konstante c4 (Zone 4)",                                11135.63, "ESt = 0,42*zvE - c4."),
    ("c5",    "Konstante c5 (Zone 5)",                                19470.38, "ESt = 0,45*zvE - c5."),
    ("soli",  "Solidaritaetszuschlag",                                   0.055, "SolzG 1995."),
    ("soli_fg", "Soli-Freigrenze (Grundtarif, festgesetzte ESt)",        20350, "2026; im Splitting doppelt. Milderungszone daneben."),
    ("soli_mz", "Milderungszone: Grenzbelastung",                        0.119, "3 Abs. 2a SolzG."),
    ("an_pausch", "Arbeitnehmer-Pauschbetrag",                             1230, "9a S. 1 Nr. 1a EStG."),
    ("sa_pausch", "Sonderausgaben-Pauschbetrag",                             36, "10c EStG."),
    ("A", "KAPITALERTRAGSTEUER UND INVESTMENTSTEUER", None, None),
    ("kapst",   "Abgeltungsteuersatz",                                    0.25, "32d Abs. 1 EStG."),
    ("kapst_eff","Effektive Belastung inkl. Soli und Kirchensteuer",     None,  "Formel: 0,25/(1+0,25*k) * (1+0,055+k) nach 32d Abs. 1 S. 3 EStG. Ohne KiSt 26,375 %, bei 9 % KiSt 27,99 %."),
    ("tfs",     "Teilfreistellung Aktienfonds (Privatvermoegen)",         0.30, "20 Abs. 1 Nr. 3 i.V.m. 20 InvStG (Aktienfondsanteil > 50 %)."),
    ("sparerpb","Sparer-Pauschbetrag (Einzelveranlagung)",               1000,  "20 Abs. 9 EStG; im Splitting 2.000 EUR."),
    ("basiszins","Basiszins Vorabpauschale 2026",                        0.032, "BMF-Schreiben vom 13.01.2026, 18 Abs. 4 InvStG (2025: 2,53 %)."),
    ("vp_faktor","Faktor fuer den Basisertrag",                           0.70, "18 Abs. 1 S. 1 InvStG."),
    ("A", "BETRIEBLICHE UND GEFOERDERTE PRIVATE ALTERSVORSORGE", None, None),
    ("st_frei_q","Steuerfrei nach 3 Nr. 63 EStG (Anteil der BBG-RV)",    0.08, "8 % der BBG-RV = 8.112 EUR in 2026."),
    ("sv_frei_q","Sozialversicherungsfrei (Anteil der BBG-RV)",           0.04, "1 Abs. 1 S. 1 Nr. 9 SvEV; 4 % = 4.056 EUR in 2026."),
    ("zul_g1",  "Zulage Stufe 1 (je EUR Eigenbeitrag)",                   0.50, "Altersvorsorgereformgesetz, in Kraft seit Mai 2026, Anwendung ab 1.1.2027."),
    ("zul_s1",  "Grenze Stufe 1",                                          360, "Bis 360 EUR Eigenbeitrag: 50 Cent je EUR."),
    ("zul_g2",  "Zulage Stufe 2 (je EUR Eigenbeitrag)",                   0.25, "Von 361 bis 1.800 EUR: 25 Cent je EUR."),
    ("zul_s2",  "Grenze Stufe 2 / Hoechstbetrag gefoerderter Eigenbeitrag", 1800, "Ergibt eine maximale Grundzulage von 540 EUR."),
    ("kind_zul","Kinderzulage je Kind (max.)",                             300, "1 EUR je eingezahltem EUR bis 300 EUR je Kind."),
    ("av_max",  "Hoechstbetrag Einzahlung Altersvorsorgevertrag p.a.",    6840, "BMF-FAQ zur Reform der geförderten privaten Altersvorsorge, Stand 05.05.2026."),
    ("av_sa_max","Hoechstbetrag Sonderausgabenabzug 10a EStG",           2340, "Eigenbeitrag bis 1.800 EUR zzgl. Zulagen. MODELLANNAHME, Vertrauensgrad ca. 70 %."),
    ("ertragsq","Ertragsanteil nicht gefoerderter Rententeile (Alter 67)", 0.17, "22 Nr. 1 S. 3 Buchst. a Doppelbuchst. bb EStG. Wird nur nachrichtlich gefuehrt."),
    ("halb",    "Steuerpflichtiger Anteil des Unterschiedsbetrags",       0.50, "20 Abs. 1 Nr. 6 S. 2 EStG analog (Vertrag > 12 Jahre, Auszahlung nach 62): halber Unterschiedsbetrag."),
    ("kv_rentner","KV-Satz auf die gesetzliche Rente (Rentneranteil)",  None,  "Halber allgemeiner Satz zzgl. halbem Zusatzbeitrag, 249a SGB V."),
    ("kv_vb",   "KV-Satz auf Versorgungsbezuege (voll)",                None,  "250 Abs. 1 Nr. 1 SGB V: der Rentner traegt den vollen Satz allein."),
]

r = 3
P = {}
for key, label, val, note in par:
    if key == "A":
        pw.cell(r, 2, label).font = H2
        pw.cell(r, 2).fill = GREYFILL
        pw.cell(r, 3).fill = GREYFILL
        r += 1
        continue
    pw.cell(r, 2, label).font = BLACK
    c = pw.cell(r, 3)
    c.border = BOX
    P[key] = f"Parameter!$C${r}"
    if val is None:
        c.font = BLACK
    else:
        c.value = val
        c.font = BLUE
        c.fill = YELLOW
        if isinstance(val, float) and val < 1.5:
            c.number_format = PCT
        elif isinstance(val, (int, float)) and abs(val) >= 1000:
            c.number_format = EUR2
        else:
            c.number_format = EUR2
    pw.cell(r, 4, note).font = Font(name=FONT, size=9, color="595959")
    r += 1

# abgeleitete Parameter
pw[P['frei_vb'].split('!')[1].replace('$', '')] = f"={P['bezug']}/20"
pw[P['frei_vb'].split('!')[1].replace('$', '')].number_format = EUR2
pw[P['kapst_eff'].split('!')[1].replace('$', '')] = f"={P['kapst']}/(1+{P['kapst']}*{IN['kist']})*(1+{P['soli']}+{IN['kist']})"
pw[P['kapst_eff'].split('!')[1].replace('$', '')].number_format = PCT
pw[P['kv_rentner'].split('!')[1].replace('$', '')] = f"=({P['kv_allg']}+{P['kv_zus']})/2"
pw[P['kv_rentner'].split('!')[1].replace('$', '')].number_format = PCT
pw[P['kv_vb'].split('!')[1].replace('$', '')] = f"={P['kv_allg']}+{P['kv_zus']}"
pw[P['kv_vb'].split('!')[1].replace('$', '')].number_format = PCT

# abgeleitete Hilfsgroessen
r += 1
pw.cell(r, 2, "ABGELEITETE GROESSEN").font = H2
pw.cell(r, 2).fill = GREYFILL
pw.cell(r, 3).fill = GREYFILL
r += 1
derived = [
    ("splitfak", "Splittingfaktor s (1 oder 2)", f"=IF({IN['stkl']}=3,2,1)", '0'),
    ("pv_an_eff", "PV-Anteil AN effektiv (inkl. Kinderlosenzuschlag)", f"={P['pv_an']}+IF({IN['kinderlos']}=1,{P['pv_kl']},0)", PCT),
    ("pv_ges_eff", "PV-Satz Rentner/Versorgungsbezuege effektiv", f"={P['pv_ges']}+IF({IN['kinderlos']}=1,{P['pv_kl']},0)", PCT),
    ("rnet_bav", "Nettorendite bAV-Depot p.a.", f"={IN['rendite']}-{IN['k_bav']}", PCT),
    ("rnet_av", "Nettorendite Altersvorsorgedepot p.a.", f"={IN['rendite']}-{IN['k_av']}", PCT),
    ("rnet_priv", "Nettorendite privates Depot p.a. (vor Steuern)", f"={IN['rendite']}-{IN['k_priv']}", PCT),
    ("n_jahre", "Jahre der Ansparphase", f"={IN['rentbeg']}-{IN['alter']}", '0'),
    ("n_ausz", "Jahre des Auszahlungsplans", f"={IN['planende']}-{IN['rentbeg']}+1", '0'),
    ("jahr_rb", "Kalenderjahr des Auszahlungsbeginns", f"=2026+{IN['rentbeg']}-{IN['alter']}", '0'),
    ("best_anteil", "Besteuerungsanteil der gesetzlichen Rente", None, PCT),
    ("sparerpb_eff", "Sparer-Pauschbetrag effektiv", f"={P['sparerpb']}*IF({IN['stkl']}=3,2,1)", EUR),
    ("wiederanl", "Wiederanlagezins nach Steuern p.a.", None, PCT),
]
for key, label, f, fmt in derived:
    pw.cell(r, 2, label).font = BLACK
    c = pw.cell(r, 3)
    c.border = BOX
    c.font = BLACK
    c.number_format = fmt
    if f:
        c.value = f
    P[key] = f"Parameter!$C${r}"
    r += 1
pw[P['best_anteil'].split('!')[1].replace('$', '')] = f"=MIN(1,0.825+0.005*({P['jahr_rb']}-2023))"
pw[P['wiederanl'].split('!')[1].replace('$', '')] = f"={P['rnet_priv']}*(1-{P['kapst_eff']}*(1-{P['tfs']}))"
pw.cell(int(P['wiederanl'].split('$')[-1]), 4,
        "Rendite, zu der ein Einmalbetrag in einem normalen Depot wieder angelegt werden koennte, "
        "nach Abgeltungsteuer und Teilfreistellung. Massstab fuer das Verrentungsaequivalent.").font = \
    Font(name=FONT, size=9, color="595959")
pw.cell(r - 2, 4, "22 Nr. 1 S. 3 Buchst. a Doppelbuchst. aa EStG: Kohorte 2023 = 82,5 %, +0,5 Prozentpunkte je Jahr bis 100 % ab 2058.").font = Font(name=FONT, size=9, color="595959")


# =====================================================================
# Hilfsfunktionen fuer Steuerformeln
# =====================================================================
def tarif(arg):
    """Grundtarif nach 32a EStG, angewandt auf die (bereits gesplittete und
    inflationsbereinigte) Bemessungsgrundlage in Zelle `arg`."""
    return (f"IF({arg}<={P['gfb']},0,"
            f"IF({arg}<={P['e2']},({P['a2']}*({arg}-{P['gfb']})/10000+{P['b2']})*({arg}-{P['gfb']})/10000,"
            f"IF({arg}<={P['e3']},({P['a3']}*({arg}-{P['e2']})/10000+{P['b3']})*({arg}-{P['e2']})/10000+{P['c3']},"
            f"IF({arg}<={P['e4']},0.42*{arg}-{P['c4']},0.45*{arg}-{P['c5']}))))")


def arg_formula(zve_ref, lam_ref):
    """Tarifargument: Splitting und Streckung des Tarifs um den Faktor lambda."""
    return (f"=MAX(0,({zve_ref}+IF({IN['stkl']}=3,{IN['partner']},0))"
            f"/{P['splitfak']}/{lam_ref})")


def est_formula(arg_ref, lam_ref):
    """Festgesetzte Einkommensteuer."""
    return f"={P['splitfak']}*{lam_ref}*({tarif(arg_ref)})"


def steuer_ges(est_ref, lam_ref):
    """ESt + Solidaritaetszuschlag (mit Freigrenze und Milderungszone) + Kirchensteuer."""
    fg = f"({P['soli_fg']}*{P['splitfak']}*{lam_ref})"
    return (f"={est_ref}+IF({est_ref}<={fg},0,"
            f"MIN({P['soli']}*{est_ref},{P['soli_mz']}*({est_ref}-{fg})))"
            f"+{IN['kist']}*{est_ref}")


# =====================================================================
# 2b) UMRECHNUNG  Nettoaufwand -> Bruttobeitrag (Bisektion)
# =====================================================================
# N(B) = B - SV-Ersparnis(B) - Steuerersparnis(B) ist stetig und streng monoton
# wachsend in B (dN/dB = 1 - Grenz-SV-Satz - Grenzsteuersatz > 0), aber wegen der
# Knicke bei 4 % und 8 % der BBG-RV, an den Beitragsbemessungsgrenzen und an den
# Tarifeckwerten nicht geschlossen invertierbar. 34 Bisektionsschritte auf
# [0; Bruttogehalt] liefern eine Genauigkeit von unter einem Zehntelcent.
uw = wb.create_sheet("Umrechnung", 3)
uw.sheet_view.showGridLines = False
uw.column_dimensions['A'].width = 3
uw.column_dimensions['B'].width = 46
uw.column_dimensions['C'].width = 16
uw.column_dimensions['D'].width = 76
uw['B1'] = "Umrechnung Nettoaufwand -> Bruttobeitrag"
uw['B1'].font = H1
uw['B2'] = ("Nur aktiv bei Vorgabemodus 2. Gesucht ist die Nullstelle von N(B) - Zielwert; N ist streng monoton "
            "wachsend, also konvergiert die Bisektion garantiert.")
uw['B2'].font = Font(name=FONT, size=10, italic=True, color="595959")

U = {}
ur = 4
konst = [
    ("tz0",    "Teilzeitfaktor im ersten Jahr",              f"=IF({IN['alter']}>={IN['tzalter']},{IN['tzfaktor']},1)", PCT1),
    ("g0",     "Bruttogehalt im ersten Jahr",                None, EUR),
    ("bbgrv0", "BBG RV im ersten Jahr",                      f"={P['bbg_rv']}", EUR),
    ("bbgkv0", "BBG KV im ersten Jahr",                      f"={P['bbg_kv']}", EUR),
    ("lam0",   "Tarifindex im ersten Jahr",                  "=1", NUM4),
    ("ziel",   "Zielwert Nettoaufwand p.a.",                 f"=12*{IN['netto_ziel']}", EUR),
    ("vors_o", "abziehbare Vorsorgeaufwendungen ohne bAV",   None, EUR),
    ("zve_o",  "zvE ohne bAV",                               None, EUR),
    ("arg_o",  "Tarifargument ohne bAV",                     None, EUR),
    ("est_o",  "ESt ohne bAV",                               None, EUR),
    ("stg_o",  "Steuer gesamt ohne bAV",                     None, EUR),
]
for key, label, fml, fmt in konst:
    uw.cell(ur, 2, label).font = BLACK
    c = uw.cell(ur, 3)
    c.font = GREEN
    c.number_format = fmt
    c.border = BOX
    if fml:
        c.value = fml
    U[key] = f"Umrechnung!$C${ur}"
    ur += 1
uw[U['g0'].split('!')[1].replace('$', '')] = f"={IN['gehalt']}*{U['tz0']}"
uw[U['vors_o'].split('!')[1].replace('$', '')] = (
    f"=MIN({U['g0']},{U['bbgrv0']})*{P['rv_an']}"
    f"+MIN({U['g0']},{U['bbgkv0']})*(0.96*({P['kv_allg']}+{P['kv_zus']})/2+{P['pv_an_eff']})")
uw[U['zve_o'].split('!')[1].replace('$', '')] = (
    f"=MAX(0,{U['g0']}-{P['an_pausch']}-{P['sa_pausch']}-{U['vors_o']})")
uw[U['arg_o'].split('!')[1].replace('$', '')] = arg_formula(U['zve_o'], U['lam0'])
uw[U['est_o'].split('!')[1].replace('$', '')] = est_formula(U['arg_o'], U['lam0'])
uw[U['stg_o'].split('!')[1].replace('$', '')] = steuer_ges(U['est_o'], U['lam0'])

UB_COLS = [("i", "Schritt", 8, '0'), ("lo", "lo", 14, EUR2), ("hi", "hi", 14, EUR2),
           ("mid", "B = (lo+hi)/2", 14, EUR2), ("svf", "sv-frei", 12, EUR2),
           ("drv", "D RV-Bem.", 12, EUR2), ("dkv", "D KV-Bem.", 12, EUR2),
           ("svan", "SV-Ersp. AN", 12, EUR2), ("svag", "SV-Ersp. AG", 12, EUR2),
           ("agfx", "AG fest", 12, EUR2),
           ("agz", "AG-Zuschuss", 12, EUR2), ("stf", "steuerfrei", 12, EUR2),
           ("vorsm", "Vorsorgeaufw.", 13, EUR2), ("zvem", "zvE mit bAV", 13, EUR2),
           ("argm", "Tarifarg.", 13, EUR2), ("estm", "ESt", 13, EUR2),
           ("stgm", "Steuer ges.", 13, EUR2), ("nb", "N(B)", 14, EUR2)]
UC = {k: gcl(i + 2) for i, (k, _, _, _) in enumerate(UB_COLS)}
UI = {k: i + 2 for i, (k, _, _, _) in enumerate(UB_COLS)}
UFMT = {k: f for k, _, _, f in UB_COLS}

UR0 = ur + 2
uw.cell(UR0 - 1, 2, "Bisektion").font = H2
for k, head, w, fmt in UB_COLS:
    uw.column_dimensions[UC[k]].width = w
    c = uw.cell(UR0, UI[k], head)
    c.font = HEADFONT
    c.fill = HEADFILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

NIT = 34
for t in range(NIT):
    r = UR0 + 1 + t
    p = r - 1
    C = {k: f"{UC[k]}{r}" for k in UC}
    Cp = {k: f"{UC[k]}{p}" for k in UC}

    def u(k, s):
        cell = uw.cell(r, UI[k], s)
        cell.number_format = UFMT[k]
        cell.font = BLACK
        return cell

    u("i", str(t + 1))
    if t == 0:
        u("lo", "=0")
        u("hi", f"={U['g0']}")
    else:
        u("lo", f"=IF({Cp['nb']}<{U['ziel']},{Cp['mid']},{Cp['lo']})")
        u("hi", f"=IF({Cp['nb']}<{U['ziel']},{Cp['hi']},{Cp['mid']})")
    u("mid", f"=({C['lo']}+{C['hi']})/2")
    u("agfx", f"=IF({C['mid']}/12+0.000000001<{IN['agz_fix_min']},0,12*{IN['agz_fix']})")
    u("svf", f"=MIN({C['mid']},MAX(0,{P['sv_frei_q']}*{U['bbgrv0']}-{C['agfx']}))")
    u("drv", f"=MIN({U['g0']},{U['bbgrv0']})-MIN({U['g0']}-{C['svf']},{U['bbgrv0']})")
    u("dkv", f"=MIN({U['g0']},{U['bbgkv0']})-MIN({U['g0']}-{C['svf']},{U['bbgkv0']})")
    u("svan", f"={C['drv']}*({P['rv_an']}+{P['av_an']})+{C['dkv']}*(({P['kv_allg']}+{P['kv_zus']})/2+{P['pv_an_eff']})")
    u("svag", f"={C['drv']}*({P['rv_an']}+{P['av_an']})+{C['dkv']}*(({P['kv_allg']}+{P['kv_zus']})/2+{P['pv_an']})")
    u("agz", f"=IF({IN['agzus_sv']}=2,MIN({IN['agzus']}*{C['mid']},{C['svag']}),"
             f"{IN['agzus']}*IF({IN['agzus_sv']}=1,{C['svf']},{C['mid']}))+{C['agfx']}")
    u("stf", f"=MIN({C['mid']},MAX(0,{P['st_frei_q']}*{U['bbgrv0']}-{C['agz']}))")
    u("vorsm", f"=MIN({U['g0']}-{C['svf']},{U['bbgrv0']})*{P['rv_an']}"
               f"+MIN({U['g0']}-{C['svf']},{U['bbgkv0']})*(0.96*({P['kv_allg']}+{P['kv_zus']})/2+{P['pv_an_eff']})")
    u("zvem", f"=MAX(0,{U['g0']}-{C['stf']}-{P['an_pausch']}-{P['sa_pausch']}-{C['vorsm']})")
    u("argm", arg_formula(C['zvem'], U['lam0']))
    u("estm", est_formula(C['argm'], U['lam0']))
    u("stgm", steuer_ges(C['estm'], U['lam0']))
    u("nb", f"={C['mid']}-{C['svan']}-({U['stg_o']}-{C['stgm']})")
    if t % 2 == 1:
        for k in UI:
            uw.cell(r, UI[k]).fill = PatternFill("solid", fgColor="F2F2F2")

ULAST = UR0 + NIT
rr = ULAST + 2
erg = [
    ("b_jahr", "ERGEBNIS: Bruttobeitrag p.a.", f"=({UC['lo']}{ULAST}+{UC['hi']}{ULAST})/2", EUR2,
     "Loesung der Gleichung N(B) = Zielwert."),
    ("b_mtl", "ERGEBNIS: Bruttobeitrag pro Monat", None, EUR2, "Wird im Blatt 'Ansparphase' verwendet, sobald Modus 2 gewaehlt ist."),
    ("resid", "Residuum N(B) - Zielwert", f"={UC['nb']}{ULAST}-{U['ziel']}", '0.000000',
     "Muss praktisch null sein. Andernfalls ist der Zielwert nicht erreichbar - siehe naechste Zeile."),
    ("warn", "Pruefung", None, None, ""),
]
for key, label, fml, fmt, note in erg:
    uw.cell(rr, 2, label).font = BOLD
    c = uw.cell(rr, 3)
    c.font = Font(name=FONT, size=11, bold=True, color="008000")
    c.border = BOX
    if fmt:
        c.number_format = fmt
    if fml:
        c.value = fml
    uw.cell(rr, 4, note).font = Font(name=FONT, size=9, color="595959")
    U[key] = f"Umrechnung!$C${rr}"
    rr += 1
uw[U['b_mtl'].split('!')[1].replace('$', '')] = f"={U['b_jahr']}/12"
uw[U['warn'].split('!')[1].replace('$', '')] = (
    f'=IF({IN["modus"]}<>2,"nicht aktiv (Modus 1)",'
    f'IF(ABS({U["resid"]})>0.01,"ZIELWERT NICHT ERREICHBAR - Nettoaufwand uebersteigt das Gehalt","ok"))')
uw[U['warn'].split('!')[1].replace('$', '')].font = Font(name=FONT, size=11, bold=True, color="C00000")

# Der ab hier gueltige effektive Monatsbeitrag
BEITRAG = f"IF({IN['modus']}=2,{U['b_mtl']},{IN['beitrag']})"

# Rueckmeldung auf dem Eingabeblatt (unterhalb der drei Ergebniszeilen)
ws.cell(res_end, 2, "Abgeleitet: Bruttobeitrag pro Monat").font = BOLD
cc = ws.cell(res_end, 3, f"={BEITRAG}")
cc.font = GREEN
cc.number_format = EUR2
ws.cell(res_end, 5, "Bei Modus 2 aus dem Nettoaufwand zurueckgerechnet, bei Modus 1 Ihre Eingabe.").font = Font(name=FONT, size=9, color="595959")
ws.cell(res_end + 1, 2, "Abgeleitet: Nettoaufwand pro Monat (erstes Jahr)").font = BOLD
cc = ws.cell(res_end + 1, 3)
cc.font = GREEN
cc.number_format = EUR2
ws.cell(res_end + 1, 5, "Der tatsaechlich aufgegebene Konsum: Bruttobeitrag minus Steuer- und SV-Ersparnis.").font = Font(name=FONT, size=9, color="595959")
NETTO_DISPLAY_CELL = ws.cell(res_end + 1, 3)

# =====================================================================
# 3) ANSPARPHASE
# =====================================================================
aw = wb.create_sheet("Ansparphase")
aw.sheet_view.showGridLines = False
aw.freeze_panes = "C8"

A_COLS = [
    ("jahr",      "Kalender-\njahr", 8, '0'),
    ("alter",     "Alter", 7, '0'),
    ("aktiv",     "erwerbs-\ntaetig", 7, '0'),
    ("anspar",    "Anspar-\nphase", 7, '0'),
    ("lohnidx",   "Lohn-\nindex", 8, NUM4),
    ("lam",       "Tarif-\nindex L", 8, NUM4),
    ("tz",        "Teilzeit-\nfaktor", 8, PCT1),
    ("brutto",    "Brutto-\ngehalt", 11, EUR),
    ("bbgrv",     "BBG RV", 11, EUR),
    ("bbgkv",     "BBG KV", 11, EUR),
    ("bav_br",    "bAV Brutto-\nbeitrag", 11, EUR),
    ("bav_stf",   "davon st.-\nfrei 3/63", 11, EUR),
    ("bav_svf",   "davon\nsv-frei", 11, EUR),
    ("agz_fx",    "AG-Zuschuss\nfest", 11, EUR),
    ("agz_var",   "AG-Zuschuss\nprozentual", 11, EUR),
    ("agz",       "AG-\nZuschuss", 11, EUR),
    ("bav_ein",   "Einzahlung\nbAV gesamt", 11, EUR),
    ("bav_gef",   "davon ge-\nfoerdert 3/63", 12, EUR),
    ("bav_ngef",  "davon NICHT\ngefoerdert", 12, EUR),
    ("drv",       "D RV-Be-\nmessung", 11, EUR),
    ("dkv",       "D KV-Be-\nmessung", 11, EUR),
    ("sverp",     "SV-Ersparnis\nAN", 11, EUR),
    ("sverp_ag",  "SV-Ersparnis\nAG", 11, EUR),
    ("vors_o",    "Vorsorgeaufw.\nohne bAV", 12, EUR),
    ("vors_m",    "Vorsorgeaufw.\nmit bAV", 12, EUR),
    ("zve_o",     "zvE ohne\nbAV", 11, EUR),
    ("zve_m",     "zvE mit\nbAV", 11, EUR),
    ("arg_o",     "Tarifarg.\nohne", 11, EUR),
    ("est_o",     "ESt ohne", 11, EUR),
    ("stg_o",     "Steuer ges.\nohne", 11, EUR),
    ("arg_m",     "Tarifarg.\nmit bAV", 11, EUR),
    ("est_m",     "ESt mit\nbAV", 11, EUR),
    ("stg_m",     "Steuer ges.\nmit bAV", 11, EUR),
    ("sterp",     "Steuer-\nersparnis", 11, EUR),
    ("netto_a",   "NETTO-\nAUFWAND", 12, EUR),
    ("av_eig",    "AV-Depot\nEigenbeitrag", 12, EUR),
    ("av_zg",     "Grund-\nzulage", 10, EUR),
    ("av_zk",     "Kinder-\nzulage", 10, EUR),
    ("av_zul",    "Zulage\ngesamt", 10, EUR),
    ("av_sa",     "Sonderausg.\n10a", 11, EUR),
    ("zve_av",    "zvE mit\nAV-SA", 11, EUR),
    ("arg_av",    "Tarifarg.\nAV", 11, EUR),
    ("est_av",    "ESt AV", 11, EUR),
    ("stg_av",    "Steuer ges.\nAV", 11, EUR),
    ("av_vort",   "Steuervorteil\ngesamt", 12, EUR),
    ("av_erst",   "zusaetzl.\nErstattung", 11, EUR),
    ("av_gef",    "Zufluss ge-\nfoerdert", 11, EUR),
    ("av_ngef",   "Zufluss nicht\ngefoerdert", 12, EUR),
    ("av_ueber",  "Ueberschuss\n> Hoechstbetr.", 12, EUR),
    ("priv_ein",  "Einzahlung\nPrivatdepot", 12, EUR),
    ("k_bav_g",   "KAPITAL bAV\ngefoerdert", 13, EUR),
    ("k_bav_n",   "KAPITAL bAV\nnicht gefoerd.", 13, EUR),
    ("k_gef",     "KAPITAL AV\ngefoerdert", 13, EUR),
    ("k_ngef",    "KAPITAL AV\nnicht gef.", 13, EUR),
    ("sl_w",      "Sleeve Wert", 12, EUR),
    ("sl_b",      "Sleeve Basis", 12, EUR),
    ("sl_vp",     "Sleeve VP", 11, EUR),
    ("sl_st",     "Sleeve VP-\nSteuer", 11, EUR),
    ("pr_w",      "KAPITAL\nPrivatdepot", 13, EUR),
    ("pr_b",      "Kostenbasis\nPrivatdepot", 12, EUR),
    ("pr_vp",     "Vorab-\npauschale", 11, EUR),
    ("pr_st",     "Steuer auf\nVorabpausch.", 12, EUR),
    ("dep",       "D Entgelt-\npunkte", 10, NUM4),
    ("dep_k",     "kumulierte\nEP-Verluste", 11, NUM4),
]
AC = {k: gcl(i + 1) for i, (k, _, _, _) in enumerate(A_COLS)}
AI = {k: i + 1 for i, (k, _, _, _) in enumerate(A_COLS)}
AFMT = {k: ff for k, _, _, ff in A_COLS}

aw['A1'] = "Ansparphase - Jahr fuer Jahr"
aw['A1'].font = H1
aw['A2'] = ("Vergleichsprinzip: gleicher NETTOAUFWAND. Der Nettoaufwand der Entgeltumwandlung (Spalte "
            + AC['netto_a'] + ") wird in den beiden anderen Optionen investiert. Der Arbeitgeberzuschuss ist "
            "das Zusatzgeld, das nur die bAV bekommt.")
aw['A2'].font = Font(name=FONT, size=9, italic=True, color="595959")

for k, head, w, fmt in A_COLS:
    col = AC[k]
    aw.column_dimensions[col].width = w
    c = aw.cell(7, AI[k], head)
    c.font = HEADFONT
    c.fill = HEADFILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
aw.row_dimensions[7].height = 34

N = 50
R0 = 8
for t in range(N):
    r = R0 + t
    p = r - 1
    first = (t == 0)

    def f(k, s):
        cell = aw.cell(r, AI[k], s)
        cell.number_format = AFMT[k]
        cell.font = BLACK
        return cell

    C = {k: f"{AC[k]}{r}" for k in AC}
    Cp = {k: f"{AC[k]}{p}" for k in AC}

    f("jahr", f"=2026+{t}")
    f("alter", f"={IN['alter']}+{t}")
    f("aktiv", f"=IF(AND({C['alter']}<{IN['ausscheid']},{C['alter']}<{IN['rentbeg']}),1,0)")
    f("anspar", f"=IF({C['alter']}<{IN['rentbeg']},1,0)")
    f("lohnidx", f"=(1+{IN['gsteig']})^{t}")
    f("lam", f"=(1+{IN['tarifanp']})^{t}")
    f("tz", f"=IF({C['alter']}>={IN['tzalter']},{IN['tzfaktor']},1)")
    f("brutto", f"={IN['gehalt']}*{C['lohnidx']}*{C['tz']}*{C['aktiv']}")
    f("bbgrv", f"={P['bbg_rv']}*{C['lohnidx']}")
    f("bbgkv", f"={P['bbg_kv']}*{C['lohnidx']}")

    # bAV
    f("bav_br", f"=MIN({C['brutto']},12*({BEITRAG})*IF({IN['dyn']}=1,{C['lohnidx']},1))*{C['aktiv']}")
    # Fester, betragsunabhaengiger Zuschuss. Faellt weg, wenn der eigene
    # Monatsbeitrag die vereinbarte Mindestschwelle nicht erreicht.
    f("agz_fx", f"={C['aktiv']}*IF({C['bav_br']}/12+0.000000001<{IN['agz_fix_min']},0,"
                f"12*{IN['agz_fix']}*IF({IN['agz_fix_dyn']}=1,{C['lohnidx']},1))")
    # Der 4-%-Topf des 1 Abs. 1 S. 1 Nr. 9 SvEV und der 8-%-Topf des 3 Nr. 63 EStG
    # werden zuerst vom Arbeitgeberbeitrag belegt, der Rest steht der Entgelt-
    # umwandlung des Arbeitnehmers zur Verfuegung.
    f("bav_svf", f"=MIN({C['bav_br']},MAX(0,{P['sv_frei_q']}*{C['bbgrv']}-{C['agz_fx']}))")
    # Modus 2 ( 1a Abs. 1a BetrAVG woertlich): Zuschuss nur bis zur tatsaechlichen
    # Ersparnis des Arbeitgebers. Oberhalb beider Beitragsbemessungsgrenzen ist sie null.
    # Der feste Zuschuss ist davon unberuehrt und kommt oben drauf.
    f("agz_var", f"=IF({IN['agzus_sv']}=2,MIN({IN['agzus']}*{C['bav_br']},{C['sverp_ag']}),"
                 f"{IN['agzus']}*IF({IN['agzus_sv']}=1,{C['bav_svf']},{C['bav_br']}))")
    f("agz", f"={C['agz_var']}+{C['agz_fx']}")
    f("bav_stf", f"=MIN({C['bav_br']},MAX(0,{P['st_frei_q']}*{C['bbgrv']}-{C['agz']}))")
    f("bav_ein", f"={C['bav_br']}+{C['agz']}")
    f("bav_gef", f"=MIN({C['bav_ein']},{P['st_frei_q']}*{C['bbgrv']})")
    f("bav_ngef", f"={C['bav_ein']}-{C['bav_gef']}")
    f("drv", f"=MIN({C['brutto']},{C['bbgrv']})-MIN({C['brutto']}-{C['bav_svf']},{C['bbgrv']})")
    f("dkv", f"=MIN({C['brutto']},{C['bbgkv']})-MIN({C['brutto']}-{C['bav_svf']},{C['bbgkv']})")
    f("sverp", f"={C['drv']}*({P['rv_an']}+{P['av_an']})+{C['dkv']}*(({P['kv_allg']}+{P['kv_zus']})/2+{P['pv_an_eff']})")
    # Arbeitgeberanteil: gleiche Haelften, aber ohne den Kinderlosenzuschlag zur Pflegeversicherung.
    f("sverp_ag", f"={C['drv']}*({P['rv_an']}+{P['av_an']})+{C['dkv']}*(({P['kv_allg']}+{P['kv_zus']})/2+{P['pv_an']})")

    # zvE
    f("vors_o", f"=MIN({C['brutto']},{C['bbgrv']})*{P['rv_an']}"
                f"+MIN({C['brutto']},{C['bbgkv']})*(0.96*({P['kv_allg']}+{P['kv_zus']})/2+{P['pv_an_eff']})")
    f("vors_m", f"=MIN({C['brutto']}-{C['bav_svf']},{C['bbgrv']})*{P['rv_an']}"
                f"+MIN({C['brutto']}-{C['bav_svf']},{C['bbgkv']})*(0.96*({P['kv_allg']}+{P['kv_zus']})/2+{P['pv_an_eff']})")
    f("zve_o", f"=MAX(0,{C['brutto']}-{P['an_pausch']}*{C['lam']}-{P['sa_pausch']}-{C['vors_o']})")
    f("zve_m", f"=MAX(0,{C['brutto']}-{C['bav_stf']}-{P['an_pausch']}*{C['lam']}-{P['sa_pausch']}-{C['vors_m']})")
    f("arg_o", arg_formula(C['zve_o'], C['lam']))
    f("est_o", est_formula(C['arg_o'], C['lam']))
    f("stg_o", steuer_ges(C['est_o'], C['lam']))
    f("arg_m", arg_formula(C['zve_m'], C['lam']))
    f("est_m", est_formula(C['arg_m'], C['lam']))
    f("stg_m", steuer_ges(C['est_m'], C['lam']))
    f("sterp", f"={C['stg_o']}-{C['stg_m']}")
    f("netto_a", f"=MAX(0,{C['bav_br']}-{C['sverp']}-{C['sterp']})")

    # Altersvorsorgedepot
    f("av_eig", f"=MIN({C['netto_a']},{P['av_max']})")
    f("av_zg", f"={P['zul_g1']}*MIN({C['av_eig']},{P['zul_s1']})"
               f"+{P['zul_g2']}*MAX(0,MIN({C['av_eig']},{P['zul_s2']})-{P['zul_s1']})")
    f("av_zk", f"=MIN({C['av_eig']},{P['kind_zul']})*{IN['kinder']}*IF({t}<{IN['kind_jahre']},1,0)")
    f("av_zul", f"=MIN({C['av_zg']}+{C['av_zk']},MAX(0,{P['av_max']}-{C['av_eig']}))")
    f("av_sa", f"=MIN(MIN({C['av_eig']},{P['zul_s2']})+{C['av_zul']},{P['av_sa_max']})")
    f("zve_av", f"=MAX(0,{C['zve_o']}-{C['av_sa']})")
    f("arg_av", arg_formula(C['zve_av'], C['lam']))
    f("est_av", est_formula(C['arg_av'], C['lam']))
    f("stg_av", steuer_ges(C['est_av'], C['lam']))
    f("av_vort", f"={C['stg_o']}-{C['stg_av']}")
    f("av_erst", f"=MAX(0,{C['av_vort']}-{C['av_zul']})")
    f("av_gef", f"=MIN({C['av_eig']},{P['zul_s2']})+{C['av_zul']}")
    f("av_ngef", f"=MAX(0,{C['av_eig']}-{P['zul_s2']})+" + ("0" if first else f"{Cp['av_erst']}"))
    f("av_ueber", f"=MAX(0,{C['netto_a']}-{P['av_max']})")
    f("priv_ein", f"={C['netto_a']}")

    # Kapitalentwicklung
    gb = f"(1+{P['rnet_bav']})"
    ga = f"(1+{P['rnet_av']})"
    gp = f"(1+{P['rnet_priv']})"
    prev_kbg = "0" if first else Cp['k_bav_g']
    f("k_bav_g", f"=IF({C['anspar']}=0,{prev_kbg},{prev_kbg}*{gb}+{C['bav_gef']}*{gb}^0.5)")
    prev_kbn = "0" if first else Cp['k_bav_n']
    f("k_bav_n", f"=IF({C['anspar']}=0,{prev_kbn},{prev_kbn}*{gb}+{C['bav_ngef']}*{gb}^0.5)")
    prev_gef = "0" if first else Cp['k_gef']
    f("k_gef", f"=IF({C['anspar']}=0,{prev_gef},{prev_gef}*{ga}+{C['av_gef']}*{ga}^0.5)")
    prev_ngef = "0" if first else Cp['k_ngef']
    f("k_ngef", f"=IF({C['anspar']}=0,{prev_ngef},{prev_ngef}*{ga}+{C['av_ngef']}*{ga}^0.5)")

    # Sleeve (Ueberschuss ueber den Hoechstbetrag; Behandlung wie Privatdepot)
    prev_slw = "0" if first else Cp['sl_w']
    prev_slb = "0" if first else Cp['sl_b']
    f("sl_vp", f"=MAX(0,MIN({prev_slw}*{gp}+{C['av_ueber']}*{gp}^0.5-{prev_slw}-{C['av_ueber']},"
               f"{prev_slw}*{P['basiszins']}*{P['vp_faktor']}))")
    f("sl_st", f"=MAX(0,{C['sl_vp']}*(1-{P['tfs']})-{P['sparerpb_eff']})*{P['kapst_eff']}")
    f("sl_w", f"=IF({C['anspar']}=0,{prev_slw},{prev_slw}*{gp}+{C['av_ueber']}*{gp}^0.5-{C['sl_st']})")
    f("sl_b", f"=IF({C['anspar']}=0,{prev_slb},{prev_slb}+{C['av_ueber']}+{C['sl_vp']})")

    # Privatdepot
    prev_prw = "0" if first else Cp['pr_w']
    prev_prb = "0" if first else Cp['pr_b']
    f("pr_vp", f"=MAX(0,MIN({prev_prw}*{gp}+{C['priv_ein']}*{gp}^0.5-{prev_prw}-{C['priv_ein']},"
               f"{prev_prw}*{P['basiszins']}*{P['vp_faktor']}))")
    f("pr_st", f"=MAX(0,{C['pr_vp']}*(1-{P['tfs']})-{P['sparerpb_eff']})*{P['kapst_eff']}")
    f("pr_w", f"=IF({C['anspar']}=0,{prev_prw},{prev_prw}*{gp}+{C['priv_ein']}*{gp}^0.5-{C['pr_st']})")
    f("pr_b", f"=IF({C['anspar']}=0,{prev_prb},{prev_prb}+{C['priv_ein']}+{C['pr_vp']})")

    # Entgeltpunkte
    f("dep", f"={C['drv']}/({P['de_rv']}*{C['lohnidx']})")
    prev_dep = "0" if first else Cp['dep_k']
    f("dep_k", f"={prev_dep}+{C['dep']}")

    if t % 2 == 1:
        for k in AI:
            aw.cell(r, AI[k]).fill = PatternFill("solid", fgColor="F2F2F2")

LAST = R0 + N - 1
NETTO_DISPLAY_CELL.value = f"=Ansparphase!${AC['netto_a']}${R0}/12"

# =====================================================================
# 4) AUSZAHLUNG
# =====================================================================
zw = wb.create_sheet("Auszahlung")
zw.sheet_view.showGridLines = False
zw.freeze_panes = "C8"

Z_COLS = [
    ("jahr", "Kalender-\njahr", 8, '0'),
    ("alter", "Alter", 7, '0'),
    ("t", "Jahr t seit\nheute", 8, '0'),
    ("lam", "Tarif-\nindex L", 8, NUM4),
    ("lohn", "Lohn-\nindex", 8, NUM4),
    ("plan", "Plan\naktiv", 6, '0'),
    ("sonst", "sonstiges\nzvE", 11, EUR),
    ("arg_b", "Tarifarg.\nBasis", 11, EUR),
    ("est_b", "ESt Basis", 11, EUR),
    ("stg_b", "Steuer ges.\nBasis", 11, EUR),
    ("bav_br_g", "bAV Zufluss\ngefoerdert", 12, EUR),
    ("bav_br_n", "bAV Zufluss\nnicht gefoerd.", 12, EUR),
    ("bav_br", "bAV Brutto-\nzufluss gesamt", 12, EUR),
    ("bav_stpfl", "davon steuer-\npflichtig", 12, EUR),
    ("bav_vb", "Versorgungs-\nbezug/Monat", 12, EUR),
    ("rente_m", "gesetzl. Rente\n/Monat", 12, EUR),
    ("vb_bpfl", "VB beitrags-\npflichtig (BBG)", 12, EUR),
    ("bav_kv", "KV auf bAV", 11, EUR),
    ("bav_pv", "PV auf bAV", 11, EUR),
    ("bav_kvpv", "KV+PV\nauf bAV", 11, EUR),
    ("mind_br", "Rentenminde-\nrung brutto", 12, EUR),
    ("mind_st", "Rentenminderung\nsteuerpflichtig", 13, EUR),
    ("mind_kv", "KV+PV-Erspar-\nnis Rente", 12, EUR),
    ("zve_bav", "zvE Option\nbAV", 11, EUR),
    ("arg_bav", "Tarifarg.\nbAV", 11, EUR),
    ("est_bav", "ESt bAV", 11, EUR),
    ("stg_bav", "Steuer ges.\nbAV", 11, EUR),
    ("zve_f", "zvE Fuenftel-\nregelung", 12, EUR),
    ("arg_f", "Tarifarg.\nFuenftel", 11, EUR),
    ("est_f", "ESt Fuenftel", 11, EUR),
    ("stg_f", "Steuer ges.\nFuenftel", 11, EUR),
    ("st_bav", "STEUER auf\nbAV (netto)", 12, EUR),
    ("bav_net", "NETTO bAV\n(nach Renten-\nminderung)", 13, EUR),
    ("gef_a", "AV gefoerdert\nStand", 12, EUR),
    ("gef_e", "Entnahme\ngefoerdert", 11, EUR),
    ("ngef_a", "AV nicht gef.\nStand", 12, EUR),
    ("ngef_e", "Entnahme\nnicht gef.", 11, EUR),
    ("sl_a", "Sleeve\nStand", 11, EUR),
    ("sl_e", "Sleeve\nEntnahme", 11, EUR),
    ("sl_st", "Sleeve\nSteuer", 11, EUR),
    ("ngef_ek", "davon Kapital\n(nicht gef.)", 12, EUR),
    ("ngef_er", "davon Rente\n(nicht gef.)", 12, EUR),
    ("av_rest", "MEMO AV-Rente\nnach dem Horizont", 14, EUR),
    ("av_stpfl", "steuerpfl.\nAV-Entnahme", 12, EUR),
    ("zve_av", "zvE Option\nAV-Depot", 11, EUR),
    ("arg_av", "Tarifarg.\nAV", 11, EUR),
    ("est_av", "ESt AV", 11, EUR),
    ("stg_av", "Steuer ges.\nAV", 11, EUR),
    ("av_net", "NETTO Alters-\nvorsorgedepot", 13, EUR),
    ("pr_a", "Privatdepot\nStand", 12, EUR),
    ("pr_ba", "Kostenbasis\nStand", 12, EUR),
    ("pr_e", "Entnahme\nPrivatdepot", 12, EUR),
    ("pr_g", "Gewinnanteil\nder Entnahme", 12, EUR),
    ("pr_st", "Abgeltung-\nsteuer", 11, EUR),
    ("pr_net", "NETTO privates\nETF-Depot", 13, EUR),
    ("df", "Diskont-\nfaktor", 9, NUM4),
    ("bav_rest", "MEMO bAV-Rente\nnach dem Horizont", 14, EUR),
    ("dfr", "Abzinsung auf\nRentenbeginn", 12, NUM4),
]
ZC = {k: gcl(i + 1) for i, (k, _, _, _) in enumerate(Z_COLS)}
ZI = {k: i + 1 for i, (k, _, _, _) in enumerate(Z_COLS)}
ZFMT = {k: ff for k, _, _, ff in Z_COLS}

zw['A1'] = "Auszahlungsphase"
zw['A1'].font = H1
zw['A2'] = ("Alle drei Optionen sind Alternativen zueinander - die Spalten NETTO ... sind daher nebeneinander, "
            "nicht additiv zu lesen. Bei der bAV ist der Verlust an gesetzlicher Rente bereits abgezogen.")
zw['A2'].font = Font(name=FONT, size=9, italic=True, color="595959")

for k, head, w, fmt in Z_COLS:
    zw.column_dimensions[ZC[k]].width = w
    c = zw.cell(7, ZI[k], head)
    c.font = HEADFONT
    c.fill = HEADFILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
zw.row_dimensions[7].height = 40

# Startkapitalien (aus der Ansparphase, Zeile LAST)
KBG = f"Ansparphase!${AC['k_bav_g']}${LAST}"
KBN = f"Ansparphase!${AC['k_bav_n']}${LAST}"
KB = f"({KBG}+{KBN})"
KG = f"Ansparphase!${AC['k_gef']}${LAST}"
KN = f"Ansparphase!${AC['k_ngef']}${LAST}"
KS = f"Ansparphase!${AC['sl_w']}${LAST}"
KSB = f"Ansparphase!${AC['sl_b']}${LAST}"
KP = f"Ansparphase!${AC['pr_w']}${LAST}"
KPB = f"Ansparphase!${AC['pr_b']}${LAST}"
EPK = f"Ansparphase!${AC['dep_k']}${LAST}"

NZ = 34
Z0 = 8
# Annuitaetenfaktoren
zw['CE4'] = "Annuitaetenfaktor Altersvorsorgedepot"
zw['CE4'].font = Font(name=FONT, size=9, color="595959")
zw['CG4'] = f"=IF({P['rnet_av']}=0,1/{P['n_ausz']},{P['rnet_av']}/(1-(1+{P['rnet_av']})^(-{P['n_ausz']})))"
zw['CG4'].number_format = PCT
zw['CE5'] = "Annuitaetenfaktor privates Depot"
zw['CE5'].font = Font(name=FONT, size=9, color="595959")
zw['CG5'] = f"=IF({P['rnet_priv']}=0,1/{P['n_ausz']},{P['rnet_priv']}/(1-(1+{P['rnet_priv']})^(-{P['n_ausz']})))"
zw['CG5'].number_format = PCT
ANN_AV = "$CG$4"
ANN_PR = "$CG$5"

for t in range(NZ):
    r = Z0 + t
    p = r - 1
    first = (t == 0)

    def g(k, s):
        cell = zw.cell(r, ZI[k], s)
        cell.number_format = ZFMT[k]
        cell.font = BLACK
        return cell

    C = {k: f"{ZC[k]}{r}" for k in ZC}
    Cp = {k: f"{ZC[k]}{p}" for k in ZC}

    g("alter", f"={IN['rentbeg']}+{t}")
    g("t", f"={C['alter']}-{IN['alter']}")
    g("jahr", f"=2026+{C['t']}")
    g("lam", f"=(1+{IN['tarifanp']})^{C['t']}")
    g("lohn", f"=(1+{IN['gsteig']})^{C['t']}")
    g("plan", f"=IF({C['alter']}<={IN['planende']},1,0)")
    g("sonst", f"={IN['sonst_zve']}*(1+{IN['infl']})^{C['t']}")
    g("arg_b", arg_formula(C['sonst'], C['lam']))
    g("est_b", est_formula(C['arg_b'], C['lam']))
    g("stg_b", steuer_ges(C['est_b'], C['lam']))

    # --- bAV
    g("df", f"=1/(1+{IN['infl']})^{C['t']}")
    g("dfr", f"=1/(1+{P['wiederanl']})^{t}")
    g("bav_rest", f"=IF(AND({IN['auszform']}=2,{C['plan']}=0),"
                  f"{KB}*{IN['rentfak']}/10000*12*(1+{IN['rentdyn']})^{t},0)")
    kap0 = "1" if first else "0"
    rente = f"{KB}*{IN['rentfak']}/10000*12*(1+{IN['rentdyn']})^{t}*{C['plan']}"
    g("bav_br_g", f"=IF({IN['auszform']}=1,IF({kap0}=1,{KBG},0),"
                  f"{rente}*IF({KB}=0,0,{KBG}/{KB}))")
    g("bav_br_n", f"=IF({IN['auszform']}=1,IF({kap0}=1,{KBN},0),"
                  f"{rente}*IF({KB}=0,0,{KBN}/{KB}))")
    g("bav_br", f"={C['bav_br_g']}+{C['bav_br_n']}")
    # Gefoerderte Schicht: voll nachgelagert ( 22 Nr. 5 S. 1 EStG).
    # Nicht gefoerderte Schicht: Kapital = halber Unterschiedsbetrag ( 22 Nr. 5 S. 2 Buchst. b
    # i.V.m. 20 Abs. 1 Nr. 6 S. 2 EStG), Rente = Ertragsanteil ( 22 Nr. 5 S. 2 Buchst. a).
    g("bav_stpfl", f"={C['bav_br_g']}+{C['bav_br_n']}*IF({IN['auszform']}=1,"
                   f"(1-IF({KBN}=0,1,MIN(1,$CG$8/{KBN})))*{P['halb']},{P['ertragsq']})")
    # Versorgungsbezug pro Monat: Kapital 1/120 fuer 120 Monate, Rente laufend
    g("bav_vb", f"=IF({IN['auszform']}=1,IF({t}<10,{KB}/120,0),{C['bav_br']}/12)")
    g("rente_m", f"={IN['rente_br']}/12*(1+{IN['gsteig']})^{C['t']}")
    # 223 Abs. 3 SGB V: Rente und Versorgungsbezuege teilen sich EINE Beitragsbemessungsgrenze.
    g("vb_bpfl", f"=MAX(0,MIN({C['bav_vb']},{P['bbg_kv']}/12*{C['lohn']}-{C['rente_m']}))")
    g("bav_kv", f"=IF({IN['kv_ruhe']}=2,0,MAX(0,{C['vb_bpfl']}-{P['frei_vb']}*{C['lohn']})*{P['kv_vb']}*12)")
    g("bav_pv", f"=IF({IN['kv_ruhe']}=2,0,IF({C['bav_vb']}>{P['frei_vb']}*{C['lohn']},{C['vb_bpfl']}*{P['pv_ges_eff']}*12,0))")
    g("bav_kvpv", f"={C['bav_kv']}+{C['bav_pv']}")

    g("mind_br", f"={EPK}*{P['rw']}*12*(1+{IN['gsteig']})^({IN['rentbeg']}-{IN['alter']})*(1+{IN['gsteig']})^{t}*{C['plan']}")
    g("mind_st", f"={C['mind_br']}*{P['best_anteil']}")
    g("mind_kv", f"=IF({IN['kv_ruhe']}=2,0,{C['mind_br']}*({P['kv_rentner']}+{P['pv_ges_eff']}))")

    g("zve_bav", f"=MAX(0,{C['sonst']}-{C['mind_st']}+{C['bav_stpfl']}-{C['bav_kvpv']}+{C['mind_kv']})")
    g("arg_bav", arg_formula(C['zve_bav'], C['lam']))
    g("est_bav", est_formula(C['arg_bav'], C['lam']))
    g("stg_bav", steuer_ges(C['est_bav'], C['lam']))
    g("zve_f", f"=MAX(0,{C['sonst']}-{C['mind_st']}+{C['bav_stpfl']}/5-{C['bav_kvpv']}+{C['mind_kv']})")
    g("arg_f", arg_formula(C['zve_f'], C['lam']))
    g("est_f", est_formula(C['arg_f'], C['lam']))
    g("stg_f", steuer_ges(C['est_f'], C['lam']))
    g("st_bav", f"=IF(AND({IN['fuenftel']}=1,{IN['auszform']}=1,{C['bav_br']}>0),"
                f"MIN({C['stg_bav']}-{C['stg_b']},5*({C['stg_f']}-{C['stg_b']})),{C['stg_bav']}-{C['stg_b']})")
    g("bav_net", f"={C['bav_br']}-{C['bav_kvpv']}-{C['st_bav']}-{C['mind_br']}+{C['mind_kv']}")

    # --- Altersvorsorgedepot
    if first:
        g("gef_a", f"={KG}")
        g("ngef_a", f"={KN}")
        g("sl_a", f"={KS}")
        g("pr_a", f"={KP}")
        g("pr_ba", f"={KPB}")
    else:
        # Bei Verrentung geht das Restkapital an den Versicherer; ein Depotkonto
        # gibt es danach nicht mehr.
        g("gef_a", f"=IF({IN['av_auszform']}=2,0,MAX(0,({Cp['gef_a']}-{Cp['gef_e']})*(1+{P['rnet_av']})))")
        g("ngef_a", f"=IF({IN['av_auszform']}=2,0,MAX(0,({Cp['ngef_a']}-{Cp['ngef_e']})*(1+{P['rnet_av']})))")
        g("sl_a", f"=MAX(0,({Cp['sl_a']}-{Cp['sl_e']})*(1+{P['rnet_priv']}))")
        g("pr_a", f"=MAX(0,({Cp['pr_a']}-{Cp['pr_e']})*(1+{P['rnet_priv']}))")
        g("pr_ba", f"=MAX(0,{Cp['pr_ba']}-{Cp['pr_ba']}*IF({Cp['pr_a']}=0,0,{Cp['pr_e']}/{Cp['pr_a']}))")

    tk = f"IF({t}=0,{IN['teilkap']},0)"
    # Verrentetes Restkapital je Schicht und insgesamt
    KGR = f"({KG}*(1-{IN['teilkap']}))"
    KNR = f"({KN}*(1-{IN['teilkap']}))"
    KRR = f"({KGR}+{KNR})"
    # Laufende Depotrente; die Aufteilung auf die Schichten folgt ihrem Anteil
    # am verrenteten Kapital, genau wie bei der bAV.
    rente_av = f"{KRR}*{IN['rentfak_av']}/10000*12*(1+{IN['rentdyn_av']})^{t}*{C['plan']}"
    g("av_rest", f"=IF(AND({IN['av_auszform']}=2,{C['plan']}=0),"
                 f"{KRR}*{IN['rentfak_av']}/10000*12*(1+{IN['rentdyn_av']})^{t},0)")
    g("gef_e", f"=IF({IN['av_auszform']}=1,"
               f"IF({C['plan']}=0,0,MIN({C['gef_a']},{KG}*{tk}+{KGR}*{ANN_AV})),"
               f"{KG}*{tk}+({rente_av})*IF({KRR}=0,0,{KGR}/{KRR}))")
    g("ngef_ek", f"=IF({IN['av_auszform']}=1,"
                 f"IF({C['plan']}=0,0,MIN({C['ngef_a']},{KN}*{tk}+{KNR}*{ANN_AV})),"
                 f"{KN}*{tk})")
    g("ngef_er", f"=IF({IN['av_auszform']}=1,0,({rente_av})*IF({KRR}=0,0,{KNR}/{KRR}))")
    g("ngef_e", f"={C['ngef_ek']}+{C['ngef_er']}")
    g("sl_e", f"=IF({C['plan']}=0,0,MIN({C['sl_a']},{KS}*{ANN_PR}))")
    g("sl_st", f"=MAX(0,{C['sl_e']}*(1-IF({KS}=0,1,MIN(1,{KSB}/{KS})))*(1-{P['tfs']}))*{P['kapst_eff']}")
    g("av_stpfl", f"={C['gef_e']}+{C['ngef_ek']}*(1-IF({KN}=0,1,MIN(1,"
                  f"Ansparphase!SUMPRODUCT_PLACEHOLDER)))*{P['halb']}+{C['ngef_er']}*{P['ertragsq']}")
    g("zve_av", f"=MAX(0,{C['sonst']}+{C['av_stpfl']})")
    g("arg_av", arg_formula(C['zve_av'], C['lam']))
    g("est_av", est_formula(C['arg_av'], C['lam']))
    g("stg_av", steuer_ges(C['est_av'], C['lam']))
    g("av_net", f"={C['gef_e']}+{C['ngef_e']}-({C['stg_av']}-{C['stg_b']})+{C['sl_e']}-{C['sl_st']}")

    # --- Privatdepot
    g("pr_e", f"=IF({C['plan']}=0,0,MIN({C['pr_a']},{KP}*{ANN_PR}))")
    g("pr_g", f"={C['pr_e']}-IF({C['pr_a']}=0,0,{C['pr_ba']}*{C['pr_e']}/{C['pr_a']})")
    g("pr_st", f"=MAX(0,{C['pr_g']}*(1-{P['tfs']})-{P['sparerpb_eff']})*{P['kapst_eff']}")
    g("pr_net", f"={C['pr_e']}-{C['pr_st']}")

    if t % 2 == 1:
        for k in ZI:
            zw.cell(r, ZI[k]).fill = PatternFill("solid", fgColor="F2F2F2")

ZLAST = Z0 + NZ - 1

# Summe der nicht gefoerderten Eigenbeitraege (fuer die Ertragsquote)
zw['CE7'] = "Summe nicht gefoerderter Eigenbeitraege (Anschaffungskosten)"
zw['CE7'].font = Font(name=FONT, size=9, color="595959")
zw['CG7'] = f"=SUM(Ansparphase!${AC['av_ngef']}${R0}:${AC['av_ngef']}${LAST})"
zw['CG7'].number_format = EUR
zw['CE8'] = "Summe nicht gefoerderter bAV-Beitraege (Anschaffungskosten)"
zw['CE8'].font = Font(name=FONT, size=9, color="595959")
zw['CG8'] = f"=SUM(Ansparphase!${AC['bav_ngef']}${R0}:${AC['bav_ngef']}${LAST})"
zw['CG8'].number_format = EUR
for t in range(NZ):
    r = Z0 + t
    zw.cell(r, ZI['av_stpfl']).value = (
        f"={ZC['gef_e']}{r}+{ZC['ngef_ek']}{r}*(1-IF({KN}=0,1,MIN(1,$CG$7/{KN})))*{P['halb']}"
        f"+{ZC['ngef_er']}{r}*{P['ertragsq']}")

# =====================================================================
# 5) CASHFLOW / IRR
# =====================================================================
cw = wb.create_sheet("Cashflow")
cw.sheet_view.showGridLines = False
cw['A1'] = "Zahlungsstroeme aus Sicht des Arbeitnehmers (fuer interne Zinsfuesse und Barwerte)"
cw['A1'].font = H1
cw['A2'] = "Negativ = Aufwand aus dem Netto, positiv = Netto-Zufluss. Alle drei Spalten haben in der Ansparphase denselben Aufwand."
cw['A2'].font = Font(name=FONT, size=9, italic=True, color="595959")
heads = ["Jahr", "Alter", "Nettoaufwand", "CF bAV (DYNO)", "CF Altersvorsorgedepot", "CF privates ETF-Depot",
         "Diskontfaktor", "BW bAV", "BW AV-Depot", "BW privat", "BW Aufwand"]
for i, h in enumerate(heads):
    c = cw.cell(4, i + 1, h)
    c.font = HEADFONT
    c.fill = HEADFILL
    c.alignment = Alignment(wrap_text=True, horizontal="center")
    cw.column_dimensions[gcl(i + 1)].width = 15
cw.row_dimensions[4].height = 30

NCF = 65
for t in range(NCF):
    r = 5 + t
    cw.cell(r, 1, f"=2026+{t}").number_format = '0'
    cw.cell(r, 2, f"={IN['alter']}+{t}").number_format = '0'
    aufw = (f"=IF({t}<{N},INDEX(Ansparphase!${AC['netto_a']}${R0}:${AC['netto_a']}${LAST},{t}+1),0)")
    cw.cell(r, 3, aufw).number_format = EUR
    for j, key in enumerate(['bav_net', 'av_net', 'pr_net']):
        col = ZC[key]
        cw.cell(r, 4 + j, f"=-$C{r}+IF(AND($B{r}>={IN['rentbeg']},$B{r}<={IN['rentbeg']}+{NZ}-1),"
                          f"INDEX(Auszahlung!${col}${Z0}:${col}${ZLAST},$B{r}-{IN['rentbeg']}+1),0)").number_format = EUR
    cw.cell(r, 7, f"=1/(1+{IN['infl']})^{t}").number_format = NUM4
    cw.cell(r, 8, f"=$G{r}*(D{r}+$C{r})").number_format = EUR
    cw.cell(r, 9, f"=$G{r}*(E{r}+$C{r})").number_format = EUR
    cw.cell(r, 10, f"=$G{r}*(F{r}+$C{r})").number_format = EUR
    cw.cell(r, 11, f"=$G{r}*$C{r}").number_format = EUR
CFLAST = 5 + NCF - 1

# =====================================================================
# 6) VERGLEICH
# =====================================================================
vw = wb.create_sheet("Vergleich")
vw.sheet_view.showGridLines = False
vw.column_dimensions['A'].width = 58
for col in "BCD":
    vw.column_dimensions[col].width = 20
vw.column_dimensions['E'].width = 86

vw['A1'] = "Ergebnisvergleich"
vw['A1'].font = H1
vw['A2'] = "Alle drei Optionen kosten in der Ansparphase exakt denselben Betrag aus dem Netto."
vw['A2'].font = Font(name=FONT, size=10, italic=True)

for i, h in enumerate(["", "bAV mit DYNO", "Altersvorsorgedepot", "privates ETF-Depot", "Anmerkung"]):
    c = vw.cell(4, i + 1, h)
    c.font = HEADFONT
    c.fill = HEADFILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

rows = [
    ("Summe eingezahlt aus dem Netto (nominal)",
     f"=SUM(Ansparphase!${AC['netto_a']}${R0}:${AC['netto_a']}${LAST})",
     "=B5", "=B5",
     "Identisch nach Konstruktion - das ist die Bedingung fuer einen fairen Vergleich."),
    ("Summe Fremdgeld (AG-Zuschuss bzw. Zulagen)",
     f"=SUM(Ansparphase!${AC['agz']}${R0}:${AC['agz']}${LAST})",
     f"=SUM(Ansparphase!${AC['av_zul']}${R0}:${AC['av_zul']}${LAST})",
     "=0",
     "Der Arbeitgeberzuschuss ist der eigentliche Grund, warum eine bAV attraktiv sein kann."),
    ("Kapital bei Auszahlungsbeginn (vor Steuern)",
     f"={KB}", f"={KG}+{KN}+{KS}", f"={KP}",
     "Beim Privatdepot ist die Vorabpauschale bereits jaehrlich abgezogen."),
    ("Summe Bruttoauszahlungen",
     f"=SUM(Auszahlung!${ZC['bav_br']}${Z0}:${ZC['bav_br']}${ZLAST})",
     f"=SUM(Auszahlung!${ZC['gef_e']}${Z0}:${ZC['gef_e']}${ZLAST})+SUM(Auszahlung!${ZC['ngef_e']}${Z0}:${ZC['ngef_e']}${ZLAST})+SUM(Auszahlung!${ZC['sl_e']}${Z0}:${ZC['sl_e']}${ZLAST})",
     f"=SUM(Auszahlung!${ZC['pr_e']}${Z0}:${ZC['pr_e']}${ZLAST})",
     "Nominal ueber den gesamten Auszahlungshorizont."),
    ("davon Einkommensteuer / Abgeltungsteuer",
     f"=SUM(Auszahlung!${ZC['st_bav']}${Z0}:${ZC['st_bav']}${ZLAST})",
     f"=SUM(Auszahlung!${ZC['stg_av']}${Z0}:${ZC['stg_av']}${ZLAST})-SUM(Auszahlung!${ZC['stg_b']}${Z0}:${ZC['stg_b']}${ZLAST})+SUM(Auszahlung!${ZC['sl_st']}${Z0}:${ZC['sl_st']}${ZLAST})",
     f"=SUM(Auszahlung!${ZC['pr_st']}${Z0}:${ZC['pr_st']}${ZLAST})",
     "bAV und Altersvorsorgedepot: nachgelagerte Besteuerung mit dem persoenlichen Grenzsteuersatz. Privatdepot: 25 % zzgl. Soli/KiSt auf 70 % des Gewinns. In der bAV-Spalte bereits saldiert mit der Steuerersparnis aus der geringeren gesetzlichen Rente."),
    ("davon Kranken- und Pflegeversicherung",
     f"=SUM(Auszahlung!${ZC['bav_kvpv']}${Z0}:${ZC['bav_kvpv']}${ZLAST})",
     "=0", "=0",
     "229 SGB V: Betriebsrenten sind Versorgungsbezuege, private Altersvorsorge nicht. In der PKV entfaellt das (Eingabe umstellen)."),
    ("Verlust an gesetzlicher Rente (brutto, kumuliert)",
     f"=SUM(Auszahlung!${ZC['mind_br']}${Z0}:${ZC['mind_br']}${ZLAST})",
     "=0", "=0",
     "Entgeltumwandlung mindert das rentenversicherungspflichtige Entgelt und damit die Entgeltpunkte."),
    ("Summe Netto-Auszahlungen (nominal)",
     f"=SUM(Auszahlung!${ZC['bav_net']}${Z0}:${ZC['bav_net']}${ZLAST})",
     f"=SUM(Auszahlung!${ZC['av_net']}${Z0}:${ZC['av_net']}${ZLAST})",
     f"=SUM(Auszahlung!${ZC['pr_net']}${Z0}:${ZC['pr_net']}${ZLAST})",
     "Nach Steuern, nach KV/PV, nach Abzug der gesetzlichen Rentenminderung."),
    ("Barwert der Netto-Auszahlungen (heutige Kaufkraft)",
     f"=SUMPRODUCT(Auszahlung!${ZC['bav_net']}${Z0}:${ZC['bav_net']}${ZLAST},Auszahlung!${ZC['df']}${Z0}:${ZC['df']}${ZLAST})",
     f"=SUMPRODUCT(Auszahlung!${ZC['av_net']}${Z0}:${ZC['av_net']}${ZLAST},Auszahlung!${ZC['df']}${Z0}:${ZC['df']}${ZLAST})",
     f"=SUMPRODUCT(Auszahlung!${ZC['pr_net']}${Z0}:${ZC['pr_net']}${ZLAST},Auszahlung!${ZC['df']}${Z0}:${ZC['df']}${ZLAST})",
     "Diskontiert mit der Inflationsrate - das ist die Groesse, die man vergleichen sollte."),
    ("Barwert des Nettoaufwands (heutige Kaufkraft)",
     f"=SUM(Cashflow!$K$5:$K${CFLAST})", "=B14", "=B14",
     "Ebenfalls identisch."),
    ("BARWERT DES NETTO-ERTRAGS (Ertrag minus Aufwand)",
     "=B13-B14", "=C13-C14", "=D13-D14",
     "Das Hauptergebnis: was jede Option in heutiger Kaufkraft ueber den eigenen Einsatz hinaus abwirft."),
    ("Verhaeltnis Ertrag zu Aufwand",
     "=IF(B14=0,0,B13/B14)", "=IF(C14=0,0,C13/C14)", "=IF(D14=0,0,D13/D14)",
     "Vielfaches des eingesetzten Nettokapitals, real."),
    ("Interner Zinsfuss (nominal, p.a.)",
     f"=IFERROR(IRR(Cashflow!$D$5:$D${CFLAST},0.05),\"n/a\")",
     f"=IFERROR(IRR(Cashflow!$E$5:$E${CFLAST},0.05),\"n/a\")",
     f"=IFERROR(IRR(Cashflow!$F$5:$F${CFLAST},0.05),\"n/a\")",
     "Rendite auf den tatsaechlichen Nettoeinsatz. Direkt mit der Bruttorendite des ETF vergleichbar."),
    ("Vorsprung gegenueber dem privaten ETF-Depot",
     "=B15-$D$15", "=C15-$D$15", "=0",
     "Positiv heisst: die gefoerderte Variante lohnt sich unter den eingestellten Annahmen."),
    ("Vorsprung in Prozent",
     "=IF($D$15=0,0,B15/$D$15-1)", "=IF($D$15=0,0,C15/$D$15-1)", "=0", ""),
]
r = 5
for lab, b, c, d, note in rows:
    vw.cell(r, 1, lab).font = BOLD if lab.isupper() or lab.startswith("BARWERT") else BLACK
    for j, fml in enumerate([b, c, d]):
        cc = vw.cell(r, 2 + j, fml)
        cc.font = GREEN
        cc.number_format = EUR
        cc.border = BOX
    vw.cell(r, 5, note).font = Font(name=FONT, size=9, color="595959")
    vw.cell(r, 5).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
for rr in (16,):
    for col in "BCD":
        vw[f"{col}{rr}"].number_format = '0.00"x"'
for rr in (17, 19):
    for col in "BCD":
        vw[f"{col}{rr}"].number_format = PCT
for col in "BCD":
    vw[f"{col}15"].font = Font(name=FONT, size=11, bold=True, color="008000")
    vw[f"{col}15"].fill = PatternFill("solid", fgColor="E2EFDA")

vw.cell(21, 1, "DIAGNOSE").font = H2
vw.cell(21, 1).fill = GREYFILL
diag = [
    ("Grenzsteuersatz in der Ansparphase (erstes Jahr)",
     f"=IF(Ansparphase!${AC['bav_stf']}${R0}=0,0,Ansparphase!${AC['sterp']}${R0}/Ansparphase!${AC['bav_stf']}${R0})",
     PCT, "Ersparnis je umgewandeltem Euro, ohne die Sozialversicherung."),
    ("Ersparnisquote gesamt (Steuer + SV) im ersten Jahr",
     f"=IF(Ansparphase!${AC['bav_br']}${R0}=0,0,1-Ansparphase!${AC['netto_a']}${R0}/Ansparphase!${AC['bav_br']}${R0})",
     PCT, "So viel vom Bruttobeitrag zahlt der Staat. Genau dieser Vorteil wird spaeter nachgelagert zurueckgeholt."),
    ("Abgabenlast der bAV-Auszahlung (Steuer + KV/PV)",
     "=IF(B8=0,0,(B9+B10)/B8)", PCT,
     "Bei Kapitalauszahlung schlaegt die Progression in einem einzigen Jahr voll durch."),
    ("Abgabenlast der Altersvorsorgedepot-Auszahlung",
     "=IF(C8=0,0,(C9+C10)/C8)", PCT, ""),
    ("Abgabenlast der Privatdepot-Auszahlung",
     "=IF(D8=0,0,(D9+D10)/D8)", PCT, "25 % zzgl. Soli auf 70 % des Gewinns, also deutlich weniger als der Grenzsteuersatz."),
    ("Tatsaechliche SV-Ersparnis des ARBEITGEBERS (erstes Jahr)",
     f"=Ansparphase!${AC['sverp_ag']}${R0}", EUR,
     "Null, wenn Ihr Entgelt auch nach der Umwandlung noch ueber beiden Beitragsbemessungsgrenzen liegt. Dann schuldet der Arbeitgeber nach 1a Abs. 1a BetrAVG auch keinen Zuschuss - er kann ihn aber freiwillig zusagen (Modus 0 oder 1)."),
    ("Tatsaechliche SV-Ersparnis des ARBEITNEHMERS (erstes Jahr)",
     f"=Ansparphase!${AC['sverp']}${R0}", EUR,
     "Ebenfalls null oberhalb der Grenzen - dafuer gehen dann auch keine Entgeltpunkte verloren. Beides haengt an derselben Groesse."),
    ("Anteil der bAV-Beitraege OHNE Foerderung (erstes Jahr)",
     f"=IF(Ansparphase!${AC['bav_ein']}${R0}=0,0,Ansparphase!${AC['bav_ngef']}${R0}/Ansparphase!${AC['bav_ein']}${R0})",
     PCT, "Alles oberhalb von 8 % der BBG-RV ist weder steuer- noch sozialabgabenbeguenstigt. Diese Schicht wandert in einen Versicherungsmantel, ohne dafuer etwas zu bekommen - ausser der Beitragspflicht in der KV. Ist dieser Anteil gross, ist die bAV strukturell unterlegen."),
    ("Arbeitgeberzuschuss prozentual (erstes Jahr)",
     f"=Ansparphase!${AC['agz_var']}${R0}", EUR,
     "Der Teil nach 1a Abs. 1a BetrAVG bzw. der vertraglich zugesagte Prozentsatz, gerechnet nach dem eingestellten Zuschussmodus."),
    ("Fester Arbeitgeberzuschuss (erstes Jahr)",
     f"=Ansparphase!${AC['agz_fx']}${R0}", EUR,
     "Betragsunabhaengiger Festzuschuss aus Tarifvertrag oder Betriebsvereinbarung. Null, wenn der eigene Monatsbeitrag die eingestellte Mindestschwelle nicht erreicht."),
    ("Fester Zuschuss oberhalb des 4-%-Topfes (erstes Jahr)",
     f"=MAX(0,Ansparphase!${AC['agz_fx']}${R0}-{P['sv_frei_q']}*Ansparphase!${AC['bbgrv']}${R0})", EUR,
     "Dieser Teil waere beim Arbeitnehmer beitragspflichtiges Arbeitsentgelt. Das Modell rechnet die zusaetzliche Beitragslast NICHT gegen und ueberschaetzt die bAV insoweit. Sollte null sein."),
    ("Ausschoepfung des Rechtsanspruchs nach 1a BetrAVG (4 % BBG)",
     f"=IF(Ansparphase!${AC['bbgrv']}${R0}=0,0,Ansparphase!${AC['bav_br']}${R0}/({P['sv_frei_q']}*Ansparphase!${AC['bbgrv']}${R0}))",
     PCT, "Ueber 100 % besteht kein Rechtsanspruch mehr auf Entgeltumwandlung; darueber hinaus ist die Zustimmung des Arbeitgebers noetig."),
    ("Versorgungsbezug pro Monat zu Beginn der Auszahlung",
     f"=Auszahlung!${ZC['bav_vb']}${Z0}", EUR2,
     "Bei Kapitalauszahlung ein Hundertzwanzigstel der Leistung fuer 120 Monate ( 229 Abs. 1 S. 3 SGB V), bei Rente die laufende Monatsrente."),
    ("Freibetrag Versorgungsbezuege im Auszahlungsjahr",
     f"={P['frei_vb']}*Auszahlung!${ZC['lohn']}${Z0}", EUR2,
     "Ein Zwanzigstel der Bezugsgroesse, 226 Abs. 2 S. 2 SGB V. 2026: 197,75 EUR; waechst mit der Lohnentwicklung."),
    ("Verhaeltnis Versorgungsbezug zu Freibetrag",
     "=IF(B32=0,0,B31/B32)", '0.00"x"',
     "Bei einem Wert bis 1,00 faellt WEDER Kranken- NOCH Pflegeversicherung an. Darueber greift die KV nur auf den uebersteigenden Teil zu, die PV wegen ihrer Freigrenze dagegen auf den GESAMTEN Betrag. Weil der Freibetrag absolut ist, faellt seine Schutzwirkung wie 1/Kapital - deshalb sinkt die Rendite der bAV streng monoton in der Beitragshoehe."),
    ("Beitragsfrei moegliches Kapital bei Kapitalauszahlung",
     "=120*B32", EUR,
     "120 x Freibetrag. Vergleichen Sie mit Zeile 7. Liegt Ihr Kapital darueber, zahlen Sie auf den Ueberschuss den vollen Beitragssatz allein."),
    ("Beitragsfrei moegliches Kapital bei lebenslanger Rente",
     f"=IF({IN['rentfak']}=0,0,B32/{IN['rentfak']}*10000)", EUR,
     "Die Verrentung streckt den Bezug auf die Restlebenszeit statt auf 120 Monate und schirmt deshalb ein Vielfaches an Kapital unter demselben Freibetrag ab. Ein Vorteil, den der Barwertvergleich ueber einen festen Horizont nicht abbildet."),
    ("MEMO: bAV-Rentenzahlungen nach dem Vergleichshorizont",
     f"=SUM(Auszahlung!${ZC['bav_rest']}${Z0}:${ZC['bav_rest']}${ZLAST})", EUR,
     "Nur bei Auszahlungsform 2. Eine lebenslange Rente zahlt ueber den Horizont hinaus - diesen Betrag beruecksichtigt der Vergleich NICHT. Wer alt wird, gewinnt hier."),
    ("MEMO: Altersvorsorgedepot-Rente nach dem Vergleichshorizont",
     f"=SUM(Auszahlung!${ZC['av_rest']}${Z0}:${ZC['av_rest']}${ZLAST})", EUR,
     "Dasselbe fuer das Altersvorsorgedepot, wenn es verrentet wird. Beide Memo-Zeilen sind der Preis eines festen Vergleichshorizonts: eine lebenslange Rente wird darin systematisch unterschaetzt."),
]
r = 22
for lab, fml, fmt, note in diag:
    vw.cell(r, 1, lab).font = BLACK
    c = vw.cell(r, 2, fml)
    c.font = GREEN
    c.number_format = fmt
    c.border = BOX
    vw.cell(r, 5, note).font = Font(name=FONT, size=9, color="595959")
    vw.cell(r, 5).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

vw.cell(38, 1, "Hinweis: Ein Foerderungsanteil deutlich ueber 0 %, eine Ausschoepfung ueber 100 % oder ein Verhaeltnis "
                "Versorgungsbezug zu Freibetrag deutlich ueber 1,00 zeigen an, dass Sie ausserhalb des beguenstigten "
                "Korridors liegen.").font = Font(name=FONT, size=9, italic=True, color="C00000")

RNG = lambda k: f"Auszahlung!${ZC[k]}${Z0}:${ZC[k]}${ZLAST}"
DFR = f"Auszahlung!${ZC['dfr']}${Z0}:${ZC['dfr']}${ZLAST}"
SP = lambda k: f"SUMPRODUCT({RNG(k)},{DFR})"
ANNW = f"({P['wiederanl']}/(1-(1+{P['wiederanl']})^(-{P['n_ausz']})))"
M = lambda expr: f"=({expr})*{ANNW}/12"

vw.cell(40, 1, "MONATLICHE LEISTUNG - alle Optionen auf denselben Zeitraum verrentet").font = H2
for cc_ in range(1, 5):
    vw.cell(40, cc_).fill = GREYFILL
for i, h in enumerate(["bAV mit DYNO", "Altersvorsorgedepot", "privates ETF-Depot"]):
    vw.cell(40, 2 + i, h).font = Font(name=FONT, size=9, bold=True, color="1F3864")
    vw.cell(40, 2 + i).alignment = Alignment(horizontal="center")

tabelle = [
    ("Wert aller Netto-Leistungen bei Rentenbeginn",
     f"={SP('bav_net')}", f"={SP('av_net')}", f"={SP('pr_net')}", EUR, False,
     "Alle Netto-Zahlungen mit dem Wiederanlagezins nach Steuern auf den Rentenbeginn bezogen. Ein Einmalbetrag und ein Zahlungsstrom werden damit vergleichbar - der Zeitpunkt wird bewertet statt ignoriert."),
    ("Bruttoleistung pro Monat",
     M(SP('bav_br')), M(f"{SP('gef_e')}+{SP('ngef_e')}+{SP('sl_e')}"), M(SP('pr_e')), EUR2, False,
     "Konstante nominale Monatszahlung ueber die Auszahlungsphase mit demselben Wert. Weil die Verrentung linear ist, addieren sich die folgenden Zeilen exakt."),
    ("abzueglich Einkommen-/Abgeltungsteuer",
     M(f"-{SP('st_bav')}"), M(f"-({SP('stg_av')}-{SP('stg_b')}+{SP('sl_st')})"), M(f"-{SP('pr_st')}"), EUR2, False,
     "bAV und Altersvorsorgedepot: persoenlicher Grenzsteuersatz. Privatdepot: nur der Gewinnanteil der Entnahme, mit 30 % Teilfreistellung."),
    ("abzueglich Kranken- und Pflegeversicherung",
     M(f"-{SP('bav_kvpv')}"), "=0", "=0", EUR2, False,
     "229, 250 SGB V. Nur die bAV ist betroffen."),
    ("abzueglich Verlust an gesetzlicher Rente",
     M(f"-({SP('mind_br')}-{SP('mind_kv')})"), "=0", "=0", EUR2, False,
     "Netto entgangene gesetzliche Rente aus der Entgeltumwandlung."),
    ("NETTO PRO MONAT (nominal)",
     M(SP('bav_net')), M(SP('av_net')), M(SP('pr_net')), EUR2, True,
     "Summe der vier Zeilen darueber. Gegenprobe fuer die Additivitaet."),
    ("NETTO PRO MONAT in heutiger Kaufkraft",
     f"=B46*Auszahlung!${ZC['df']}${Z0}", f"=C46*Auszahlung!${ZC['df']}${Z0}",
     f"=D46*Auszahlung!${ZC['df']}${Z0}", EUR2, True,
     "Die einzige Zahl, die man unmittelbar mit dem heutigen Nettoeinkommen vergleichen kann."),
]
r = 41
for lab, b, c, d, fmt, hl, note in tabelle:
    vw.cell(r, 1, lab).font = BOLD if hl else BLACK
    for j, fml in enumerate([b, c, d]):
        cc = vw.cell(r, 2 + j, fml)
        cc.number_format = fmt
        cc.border = BOX
        if hl:
            cc.font = Font(name=FONT, size=11, bold=True, color="008000")
            cc.fill = PatternFill("solid", fgColor="E2EFDA")
        else:
            cc.font = GREEN
    vw.cell(r, 5, note).font = Font(name=FONT, size=9, color="595959")
    vw.cell(r, 5).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
vw.cell(48, 1, "Der Bezugszeitraum ist fuer alle drei Optionen identisch (Rentenbeginn bis Vergleichshorizont). "
               "Bei Kapitalauszahlung wird der Einmalbetrag also nicht durch zwoelf geteilt, sondern mit dem "
               "Wiederanlagezins aus dem Blatt 'Parameter' auf denselben Zeitraum verrentet. Wer ihn verkonsumiert "
               "statt anzulegen, steht schlechter - dann den Wiederanlagezins auf 0 setzen.").font = \
    Font(name=FONT, size=9, italic=True, color="595959")
vw.cell(48, 1).alignment = Alignment(wrap_text=True, vertical="top")
vw.row_dimensions[48].height = 26

vw['A50'] = "Barwert Netto-Ertrag bAV"
vw['B50'] = "=B15"
vw['A51'] = "Barwert Netto-Ertrag Altersvorsorgedepot"
vw['B51'] = "=C15"
vw['A52'] = "Barwert Netto-Ertrag privates Depot"
vw['B52'] = "=D15"
for rr in (50, 51, 52):
    vw[f"B{rr}"].number_format = EUR
    vw[f"A{rr}"].font = Font(name=FONT, size=9, color="808080")
    vw[f"B{rr}"].font = Font(name=FONT, size=9, color="808080")

ch = BarChart()
ch.type = "col"
ch.title = "Barwert des Netto-Ertrags (heutige Kaufkraft)"
ch.y_axis.title = "EUR"
data = Reference(vw, min_col=2, min_row=50, max_row=52)
cats = Reference(vw, min_col=1, min_row=50, max_row=52)
ch.add_data(data, titles_from_data=False)
ch.set_categories(cats)
ch.legend = None
ch.height = 8
ch.width = 17
vw.add_chart(ch, "A55")

# =====================================================================
# 7) ANLEITUNG
# =====================================================================
dw = wb.create_sheet("Anleitung", 0)
dw.sheet_view.showGridLines = False
dw.column_dimensions['A'].width = 4
dw.column_dimensions['B'].width = 130
txt = [
    ("H1", "bAV mit DYNO  vs.  Altersvorsorgedepot  vs.  privates ETF-Depot"),
    ("i", "Rechtsstand 2026. Alle Zahlen sind Formeln - aendern Sie die gelben Zellen im Blatt 'Eingaben' und alles rechnet neu."),
    ("", ""),
    ("H2", "Aufbau"),
    ("", "Begriffe       - Legende aller Fachbegriffe und Spaltenueberschriften. Bei Unklarheiten zuerst dort nachsehen."),
    ("", "Eingaben       - Ihre persoenlichen Groessen. Nur hier tippen."),
    ("", "Umrechnung     - Bisektion, die aus einem vorgegebenen Nettoaufwand den noetigen Bruttobeitrag bestimmt."),
    ("", "Parameter      - gesetzliche Rechengroessen 2026 mit Fundstelle. Anpassen, wenn sich das Recht aendert."),
    ("", "Ansparphase    - Jahr fuer Jahr: Brutto, SV, Steuer, Nettoaufwand, Einzahlungen, Kapitalentwicklung, Entgeltpunkte."),
    ("", "Auszahlung     - Jahr fuer Jahr: Bruttozufluss, KV/PV, Steuer, Netto je Option."),
    ("", "Cashflow       - Zahlungsstroeme fuer interne Zinsfuesse und Barwerte."),
    ("", "Vergleich      - das Ergebnis nebeneinander, mit Diagramm."),
    ("", ""),
    ("H2", "Das Vergleichsprinzip"),
    ("", "Ein Vergleich von Bruttobeitraegen waere unfair: die Entgeltumwandlung kommt aus dem Brutto, das private Sparen aus dem Netto."),
    ("", "Deshalb wird der NETTOAUFWAND gleichgesetzt. Aus dem eingegebenen Bruttobeitrag folgt, wieviel Netto tatsaechlich fehlt"),
    ("", "(Beitrag minus gesparte Lohnsteuer minus gesparter Arbeitnehmeranteil zur Sozialversicherung). Genau dieser Betrag wird in"),
    ("", "den beiden anderen Optionen investiert. Was uebrig bleibt, ist ein echter Vergleich der Foerderarchitekturen."),
    ("", "Ueber den Vorgabemodus koennen Sie die Richtung umkehren und direkt den Nettoaufwand festlegen - oekonomisch ist"),
    ("", "das die natuerlichere Groesse, denn sie misst den tatsaechlich aufgegebenen Konsum. Der Bruttobeitrag des ersten"),
    ("", "Jahres wird dann zurueckgerechnet und danach nach der eingestellten Dynamisierungsregel fortgeschrieben; in"),
    ("", "spaeteren Jahren kann der Nettoaufwand daher leicht abweichen (Teilzeit, Tarifprogression, Beitragsbemessungsgrenzen)."),
    ("", ""),
    ("H2", "Die drei Optionen im Kern"),
    ("", "1) bAV (Direktversicherung, ETF-basiert, z.B. DYNO)"),
    ("", "   + Beitrag steuerfrei bis 8 % der BBG-RV ( 3 Nr. 63 EStG), sozialabgabenfrei bis 4 % ( 1 Abs. 1 S. 1 Nr. 9 SvEV)"),
    ("", "   + Arbeitgeberzuschuss von mindestens 15 % ( 1a Abs. 1a BetrAVG) - das ist der eigentliche Renditetreiber"),
    ("", "   + ggf. zusaetzlich ein tarifvertraglicher FESTBETRAG unabhaengig vom Monatsbeitrag (z.B. Deutsche Bahn)"),
    ("", "   - Leistung voll steuerpflichtig ( 22 Nr. 5 S. 1 EStG); Kapitalabfindung trifft in einem Jahr den Spitzensteuersatz"),
    ("", "   - Kranken- und Pflegeversicherung auf die volle Leistung, Beitragssatz allein vom Rentner ( 229, 250 SGB V)"),
    ("", "   - Entgeltumwandlung mindert die Entgeltpunkte und damit die gesetzliche Rente"),
    ("", "   - Produktkosten des Versicherungsmantels, gebundenes Kapital bis zum Rentenalter"),
    ("", "2) Altersvorsorgedepot (ab 1.1.2027, Altersvorsorgereformgesetz)"),
    ("", "   + Zulage 50 Cent je Euro bis 360 EUR, 25 Cent je Euro bis 1.800 EUR (max. 540 EUR Grundzulage), 300 EUR je Kind"),
    ("", "   + steuerfreie Thesaurierung in der Ansparphase - keine Vorabpauschale, kein Steuerabzug auf Ausschuettungen"),
    ("", "   + KEINE Kranken- und Pflegeversicherungsbeitraege auf die Leistung"),
    ("", "   - gefoerderte Teile voll nachgelagert steuerpflichtig; Beitraege ueber 1.800 EUR sind nicht gefoerdert"),
    ("", "   - Auszahlung fruehestens ab 65, Kapital gebunden, Anlageuniversum durch Positivliste begrenzt"),
    ("", "3) Privates ETF-Depot"),
    ("", "   + voellig flexibel, jederzeit verfuegbar, vererbbar, keine Produktbindung, niedrigste Kosten"),
    ("", "   - Vorabpauschale nach 18 InvStG in jedem Jahr mit Wertzuwachs (Basiszins 2026: 3,20 %)"),
    ("", "   - Abgeltungsteuer 25 % zzgl. Soli und ggf. Kirchensteuer auf 70 % des Gewinns (30 % Teilfreistellung)"),
    ("", "   - keine Foerderung, kein Fremdgeld"),
    ("", ""),
    ("H2", "Wichtige Modellannahmen (bitte lesen, bevor Sie dem Ergebnis glauben)"),
    ("", "- Das zu versteuernde Einkommen wird aus dem Brutto naeherungsweise abgeleitet: abzueglich Arbeitnehmer-Pauschbetrag,"),
    ("", "  Sonderausgaben-Pauschbetrag und der abziehbaren Vorsorgeaufwendungen (RV voll, KV zu 96 %, PV voll). Genauigkeit ca. 1 %."),
    ("", "- Der Einkommensteuertarif wird mit dem eingegebenen Satz gestreckt: ESt = s * L * T(zvE / s / L). Das ist exakt aequivalent"),
    ("", "  zu einer proportionalen Verschiebung aller Tarifeckwerte. Auf 0 % setzen, um die kalte Progression zu sehen."),
    ("", "- BBG, Bezugsgroesse, Durchschnittsentgelt und Rentenwert werden mit der Lohnsteigerung fortgeschrieben."),
    ("", "- Oberhalb der Beitragsbemessungsgrenzen ist die Entgeltumwandlung sozialversicherungsrechtlich neutral: die"),
    ("", "  Bemessungsgrundlage bleibt gedeckelt, also spart weder Arbeitnehmer noch Arbeitgeber Beitraege, und es gehen"),
    ("", "  auch keine Entgeltpunkte verloren. Beides folgt aus derselben Differenz und ist im Modell zwingend gekoppelt."),
    ("", "  Der Zuschussmodus 2 zieht daraus die Konsequenz und setzt den prozentualen Zuschuss dann auf null."),
    ("", "- Der FESTE Arbeitgeberzuschuss (z.B. Deutsche Bahn) ist betragsunabhaengig und vom Zuschussmodus unberuehrt: er"),
    ("", "  beruht auf Tarifvertrag oder Betriebsvereinbarung, nicht auf 1a Abs. 1a BetrAVG, und kommt oben drauf."),
    ("", "  Als arbeitgeberfinanzierte Zuwendung nach 3 Nr. 63 EStG belegt er den 8-%-Topf und die 4-%-Beitragsfreiheit"),
    ("", "  VORRANGIG; die eigene Entgeltumwandlung wird insoweit verdraengt und teurer. Sein Barwert waechst nicht mit dem"),
    ("", "  Beitrag - oekonomisch ist er ein Sockel, kein Hebel: er hebt den Durchschnittsertrag, nicht die Grenzrendite."),
    ("", "  NICHT abgebildet: uebersteigt der feste Zuschuss allein 4 % der BBG-RV, waere der Ueberhang beim Arbeitnehmer"),
    ("", "  beitragspflichtiges Arbeitsentgelt. Das Modell rechnet diese Beitragslast nicht gegen und ueberschaetzt die bAV"),
    ("", "  insoweit; das Blatt 'Kennzahlen' weist den Ueberhang aus. Vertrauensgrad der Vorrang-Lesart ca. 80 %."),
    ("", "- Beitraege gelten als unterjaehrig geleistet (Verzinsung mit (1+r)^0,5 im Einzahlungsjahr)."),
    ("", "- In der Auszahlungsphase wird die Vorabpauschale nicht mehr fortgeschrieben; sie waere durch die Anrechnung auf den"),
    ("", "  realisierten Gewinn ( 19 InvStG) barwertig weitgehend neutral."),
    ("", "- Der Sonderausgabenhoechstbetrag von 2.340 EUR fuer das Altersvorsorgedepot ist eine Modellannahme aus der BMF-FAQ"),
    ("", "  (Eigenbeitrag bis 1.800 EUR zuzueglich Zulagen). Vertrauensgrad rund 70 % - im Blatt 'Parameter' aenderbar."),
    ("", "- Guenstigerpruefung nach 10a Abs. 2 EStG: die Gesamtfoerderung ist das MAXIMUM aus Zulagenanspruch und Steuervorteil"),
    ("", "  aus dem Sonderausgabenabzug, nicht deren Summe. Solange der Steuervorteil groesser ist, erhoeht eine zusaetzliche"),
    ("", "  Kinderzulage die Foerderung um exakt null - sie wird vollstaendig angerechnet. Erst wenn die Zulage den Steuervorteil"),
    ("", "  uebersteigt, fliesst echtes zusaetzliches Geld. Das ist der Grund fuer scheinbar paradoxe Ergebnisse bei einem Kind."),
    ("", "- Die Steuererstattung aus der Guenstigerpruefung wird im Folgejahr als NICHT gefoerderter Beitrag in denselben Vertrag"),
    ("", "  reinvestiert. Das ist eine Konvention, keine Rechtsfolge; sie haelt den Nettoaufwand aller drei Optionen exakt gleich."),
    ("", "- Die bAV wird in ZWEI Schichten gefuehrt: gefoerdert (bis 8 % der BBG-RV, 3 Nr. 63 EStG, voll nachgelagert"),
    ("", "  steuerpflichtig) und NICHT gefoerdert (der Rest, aus versteuertem Entgelt; bei Kapitalauszahlung nur halber"),
    ("", "  Unterschiedsbetrag nach 22 Nr. 5 S. 2 Buchst. b i.V.m. 20 Abs. 1 Nr. 6 S. 2 EStG, bei Rente Ertragsanteil)."),
    ("", "  Ohne diese Trennung wuerde derselbe Euro zweimal besteuert - und der Barwert der bAV bei hohen Beitraegen negativ."),
    ("", "- Die Beitragsbemessungsgrenze der KV wird auf die Summe aus gesetzlicher Rente und Versorgungsbezuegen angewandt"),
    ("", "  ( 223 Abs. 3 SGB V). Dafuer wird die gesetzliche Bruttorente als Eingabe benoetigt."),
    ("", "- Bei der lebenslangen bAV-Rente ist der Vergleich horizontabhaengig: das Modell zaehlt nur die Zahlungen bis zum"),
    ("", "  eingestellten Vergleichshorizont. Die Zahlungen danach stehen als Memo-Zeile im Blatt 'Vergleich'. Eine Leibrente ist"),
    ("", "  primaer eine Versicherung gegen Langlebigkeit, kein Renditeprodukt - sie kann deshalb im Barwertvergleich nur verlieren."),
    ("", "- Nicht modelliert: Abschlaege bei vorzeitigem Rentenbezug, Auswirkungen der Entgeltumwandlung auf Kranken-, Arbeitslosen-"),
    ("", "  und Elterngeld, Hinterbliebenenversorgung, Insolvenzsicherung (PSVaG), Portabilitaet bei Arbeitgeberwechsel,"),
    ("", "  Kinderabschlaege in der Pflegeversicherung, Guenstigerpruefung nach 32d Abs. 6 EStG."),
    ("", ""),
    ("H2", "Quellen"),
    ("", "SVBezGrV 2026 - https://www.gesetze-im-internet.de/svbezgrv_2026/BJNR1160A0025.html"),
    ("", "32a EStG - https://www.gesetze-im-internet.de/estg/__32a.html"),
    ("", "BMF, FAQ zur Reform der gefoerderten privaten Altersvorsorge, 05.05.2026 -"),
    ("", "   https://www.bundesfinanzministerium.de/Content/DE/FAQ/reform-der-privaten-altersvorsorge.html"),
    ("", "Bundesregierung, Private Altersvorsorge wird attraktiver, 01.06.2026 -"),
    ("", "   https://www.bundesregierung.de/breg-de/aktuelles/reform-private-altersvorsorge-2400072"),
    ("", "BMF-Schreiben vom 13.01.2026, Basiszins zur Vorabpauschale ( 18 Abs. 4 InvStG) -"),
    ("", "   https://www.bundesfinanzministerium.de/Content/DE/Downloads/BMF_Schreiben/Steuerarten/Investmentsteuer/2026-01-13-basiszins-berechnung-vorabpauschale.pdf"),
    ("", "GKV-Spitzenverband, Rechengroessen 2026 - https://www.gkv-spitzenverband.de/media/dokumente/presse/zahlen_und_grafiken/20260101_Faktenblatt_Rechengroessen_Beitragsrecht.pdf"),
    ("", "DYNO - https://heydyno.de"),
    ("", ""),
    ("i", "Keine Steuer- oder Anlageberatung. Ich bin weder Steuerberater noch Anlageberater; das Modell ersetzt keine Pruefung des konkreten Vertrags."),
]
r = 2
for style, line in txt:
    c = dw.cell(r, 2, line)
    if style == "H1":
        c.font = H1
    elif style == "H2":
        c.font = H2
    elif style == "i":
        c.font = Font(name=FONT, size=10, italic=True, color="595959")
    else:
        c.font = Font(name=FONT, size=10)
    r += 1

# =====================================================================
# 8) BEGRIFFE
# =====================================================================
bw = wb.create_sheet("Begriffe", 1)
bw.sheet_view.showGridLines = False
bw.column_dimensions['A'].width = 3
bw.column_dimensions['B'].width = 34
bw.column_dimensions['C'].width = 104
bw.column_dimensions['D'].width = 44
bw['B1'] = "Begriffslegende"
bw['B1'].font = H1
bw['B2'] = ("Die zentralen Groessen des Modells. Wo eine Rechtsnorm existiert, steht sie in der dritten Spalte; "
            "wo es eine Modellkonvention ist, steht das ausdruecklich dort.")
bw['B2'].font = Font(name=FONT, size=10, italic=True, color="595959")

GLOSS = [
    ("H", "Die Kerngroesse des Vergleichs", None, None),
    ("Nettoaufwand", (
        "Betrag, um den Ihr verfuegbares Einkommen im Beitragsjahr tatsaechlich sinkt:\n"
        "    Nettoaufwand  =  Bruttobeitrag  -  Lohnsteuerersparnis  -  ersparter AN-Anteil zur Sozialversicherung\n"
        "Beispiel Basisfall: 3.600 - 1.237 - 382 = 1.981 EUR. Sie wandeln 3.600 EUR Brutto um, spueren aber nur 1.981 EUR.\n"
        "Genau dieser Betrag - nicht der Bruttobeitrag - wird in den beiden anderen Optionen investiert, denn privates "
        "Sparen erfolgt aus versteuertem Einkommen. Ohne diese Normierung wuerde man 3.600 EUR brutto gegen 3.600 EUR "
        "netto vergleichen und die bAV systematisch bevorzugen."),
     "Spalte AE im Blatt 'Ansparphase'"),
    ("Vorgabemodus", (
        "Modus 1: Sie geben den Bruttobeitrag vor, der Nettoaufwand ergibt sich.\n"
        "Modus 2: Sie geben den Nettoaufwand vor - also den aufgegebenen Konsum - und der noetige Bruttobeitrag wird "
        "zurueckgerechnet.\n"
        "Die Umkehrung ist nicht geschlossen darstellbar, weil N(B) Knicke bei 4 % und 8 % der BBG-RV, an den "
        "Beitragsbemessungsgrenzen und an den Tarifeckwerten hat. N ist aber stetig und streng monoton wachsend "
        "(dN/dB = 1 - Grenz-SV-Satz - Grenzsteuersatz, im Regelfall etwa 0,5 bis 0,6), also konvergiert eine Bisektion "
        "garantiert. Das Blatt 'Umrechnung' fuehrt 34 Schritte auf dem Intervall [0; Bruttogehalt] aus."),
     "Blatt 'Umrechnung'"),
    ("Fremdgeld", (
        "Geld, das nicht von Ihnen kommt: der Arbeitgeberzuschuss bei der bAV, die Zulagen beim Altersvorsorgedepot. "
        "Der einzige Posten, der eine gefoerderte Anlage strukturell besser machen kann als ein freies Depot - "
        "der Steueraufschub allein leistet das nicht."),
     "Zeile 6 im Blatt 'Vergleich'"),

    ("H", "Steuer", None, None),
    ("Grenzsteuersatz", (
        "Steuer auf den naechsten verdienten Euro, also die Ableitung des Tarifs. Entscheidend fuer den Wert eines "
        "Abzugs. 2026 zwischen 14 % und 45 %."),
     "32a EStG"),
    ("Durchschnittssteuersatz", (
        "Gesamte Steuer geteilt durch das gesamte zu versteuernde Einkommen. Immer kleiner als der Grenzsteuersatz. "
        "Relevant fuer die Auszahlungsphase, wenn eine grosse Summe auf einmal zufliesst."), None),
    ("zvE", (
        "Zu versteuerndes Einkommen: Bruttogehalt abzueglich Werbungskosten (mindestens Arbeitnehmer-Pauschbetrag), "
        "Sonderausgaben-Pauschbetrag und abziehbarer Vorsorgeaufwendungen. Bemessungsgrundlage des Tarifs."),
     "2 Abs. 5 EStG"),
    ("Splittingtarif", (
        "Bei Zusammenveranlagung: ESt = 2 x T((zvE1 + zvE2)/2). Senkt den Grenzsteuersatz stark, solange das halbierte "
        "Einkommen in der Progressionszone liegt, und wirkt gar nicht mehr im Spitzensteuersatz."),
     "32a Abs. 5 EStG"),
    ("Tarifindex L (lambda)", (
        "Streckungsfaktor des Steuertarifs fuer kuenftige Jahre. Das Modell rechnet ESt = s * L * T(zvE / s / L), was "
        "exakt einer proportionalen Verschiebung aller Tarifeckwerte entspricht. L = 1 bedeutet: keine Anpassung, also "
        "volle kalte Progression."), "Modellkonvention"),
    ("Nachgelagerte Besteuerung", (
        "Beitraege bleiben steuerfrei, dafuer ist die spaetere Leistung voll steuerpflichtig. Rechnerisch ist der Aufschub "
        "fuer sich genommen wertneutral, weil (1-t) * (1+r)^n = (1+r)^n * (1-t). Er lohnt nur, wenn der Steuersatz bei "
        "Auszahlung niedriger ist als bei Einzahlung."), "22 Nr. 5 S. 1 EStG"),
    ("Fuenftelregelung", (
        "Ermaessigte Besteuerung ausserordentlicher Einkuenfte: die Steuer wird auf ein Fuenftel berechnet und "
        "verfuenffacht. Bei planmaessigen Kapitalabfindungen aus 3 Nr. 63-Vertraegen nach der Rechtsprechung nur "
        "ausnahmsweise anwendbar - im Modell abschaltbar und per Default aus."), "34 EStG"),

    ("H", "Sozialversicherung", None, None),
    ("Beitragsbemessungsgrenze (BBG)", (
        "Obergrenze des beitragspflichtigen Entgelts. 2026: 101.400 EUR in RV und AV, 69.750 EUR in KV und PV. "
        "Oberhalb der Grenze ist eine Entgeltumwandlung sozialversicherungsrechtlich neutral - weder Sie noch Ihr "
        "Arbeitgeber sparen etwas, und es gehen auch keine Entgeltpunkte verloren."), "159 ff. SGB VI, 223 SGB V"),
    ("Bezugsgroesse", (
        "Rechengroesse des Sozialrechts, 2026: 3.955 EUR monatlich. Wichtig hier vor allem als Basis des Freibetrags "
        "auf Versorgungsbezuege (ein Zwanzigstel)."), "18 SGB IV"),
    ("Versorgungsbezug", (
        "Leistung aus betrieblicher Altersversorgung. In der GKV beitragspflichtig, und zwar mit dem VOLLEN Beitragssatz, "
        "den der Rentner allein traegt. Eine Kapitalabfindung wird dafuer auf 120 Monate verteilt. Leistungen aus "
        "PRIVATER Altersvorsorge sind keine Versorgungsbezuege und bleiben beitragsfrei - der groesste einzelne "
        "Unterschied zwischen den drei Optionen."), "229, 250 Abs. 1 Nr. 1 SGB V"),
    ("Freibetrag / Freigrenze", (
        "Beim Freibetrag bleibt nur der Betrag bis zur Grenze unbelastet, der Rest wird belastet (KV: 197,75 EUR/Monat "
        "in 2026). Bei der Freigrenze wird bei Ueberschreiten der GESAMTE Betrag belastet - so in der Pflegeversicherung. "
        "Mathematisch: stetige gegen unstetige Funktion."), "226 Abs. 2 SGB V"),
    ("Entgeltpunkt", (
        "Beitragspflichtiges Entgelt geteilt durch das Durchschnittsentgelt aller Versicherten (2026: 51.944 EUR). "
        "Ein Punkt ergibt seit Juli 2026 eine Monatsrente von 42,52 EUR. Entgeltumwandlung senkt das beitragspflichtige "
        "Entgelt und damit die Punkte - aber nur unterhalb der BBG."), "63, 70 SGB VI"),
    ("Besteuerungsanteil", (
        "Anteil der gesetzlichen Rente, der steuerpflichtig ist. Haengt vom Jahr des Rentenbeginns ab: 82,5 % fuer die "
        "Kohorte 2023, danach +0,5 Prozentpunkte pro Jahr bis 100 % ab 2058."),
     "22 Nr. 1 S. 3 Buchst. a Doppelbuchst. aa EStG"),

    ("H", "Betriebliche Altersvorsorge", None, None),
    ("Entgeltumwandlung", (
        "Umwandlung kuenftiger Gehaltsansprueche in eine Anwartschaft auf Betriebsrente. Ein Rechtsanspruch besteht nur "
        "bis 4 % der BBG-RV; darueber hinaus braucht es die Zustimmung des Arbeitgebers."), "1a BetrAVG"),
    ("Gefoerderte Schicht", (
        "Der Teil der Beitraege bis 8 % der BBG-RV (2026: 8.112 EUR). Steuerfrei beim Einzahlen, dafuer die Leistung "
        "spaeter zu 100 % steuerpflichtig."), "3 Nr. 63 EStG"),
    ("Nicht gefoerderte Schicht", (
        "Alles darueber. Fliesst aus versteuertem und verbeitragtem Entgelt in den Vertrag. Deshalb ist die Leistung "
        "spaeter NICHT voll steuerpflichtig, sondern nur mit dem halben Unterschiedsbetrag (Kapital) bzw. dem "
        "Ertragsanteil (Rente). Die KV-Pflicht trifft sie trotzdem in voller Hoehe."),
     "22 Nr. 5 S. 2 EStG"),
    ("Unterschiedsbetrag", (
        "Auszahlung minus eingezahlte Beitraege, also der reine Wertzuwachs. Bei Vertragsdauer ueber zwoelf Jahren und "
        "Auszahlung nach dem 62. Lebensjahr ist davon nur die Haelfte steuerpflichtig."), "20 Abs. 1 Nr. 6 S. 2 EStG"),
    ("Ertragsanteil", (
        "Pauschal unterstellter Zinsanteil einer Leibrente aus nicht gefoerdertem Kapital, abhaengig vom Alter bei "
        "Rentenbeginn (bei 67: 17 %). Nur dieser Anteil wird besteuert."),
     "22 Nr. 1 S. 3 Buchst. a Doppelbuchst. bb EStG"),
    ("Rentenfaktor", (
        "Monatliche Rente je 10.000 EUR Kapital. Ein Faktor von 26 entspricht etwa 3,1 % Entnahmequote pro Jahr - der "
        "Versicherer kalkuliert fuer ein sehr hohes Endalter. Eine Leibrente ist primaer eine Versicherung gegen "
        "Langlebigkeit, kein Renditeprodukt; im Barwertvergleich ueber einen festen Horizont kann sie deshalb nur "
        "verlieren."), "Vertragsbedingung"),

    ("H", "Gefoerderte private Altersvorsorge (Altersvorsorgedepot)", None, None),
    ("Zulage", (
        "Direkter staatlicher Zuschuss: 50 Cent je Euro bis 360 EUR Eigenbeitrag, 25 Cent je Euro bis 1.800 EUR, "
        "also hoechstens 540 EUR Grundzulage. Zusaetzlich 1 Euro je Euro bis 300 EUR je Kind, solange Kindergeld "
        "zusteht."), "Abschnitt XI EStG i.d.F. des Altersvorsorgereformgesetzes"),
    ("Sonderausgabenabzug", (
        "Alternative Foerderung: Eigenbeitraege bis 1.800 EUR zuzueglich Zulagen mindern das zvE."), "10a Abs. 1 EStG"),
    ("Guenstigerpruefung", (
        "Die Gesamtfoerderung ist das MAXIMUM aus Zulagenanspruch und Steuervorteil des Sonderausgabenabzugs, nicht "
        "deren Summe: F = max(Zulage, Steuervorteil). Solange der Steuervorteil groesser ist, erhoeht eine zusaetzliche "
        "Zulage die Foerderung um exakt null - sie wird vollstaendig angerechnet. Deshalb kann ein einzelnes Kind das "
        "Ergebnis sogar minimal verschlechtern, waehrend zwei Kinder es deutlich verbessern."), "10a Abs. 2 EStG"),
    ("Standarddepot", (
        "Kostenreguliertes Standardprodukt, das jeder Anbieter fuehren muss. Die Effektivkosten sind auf 1,0 % p.a. "
        "begrenzt."), "Altersvorsorgereformgesetz"),

    ("H", "Privates Depot", None, None),
    ("Vorabpauschale", (
        "Jaehrliche Vorwegbesteuerung thesaurierender Fonds. Bemessungsgrundlage ist der Basisertrag = Wert am "
        "Jahresanfang x Basiszins x 70 %, gedeckelt auf den tatsaechlichen Wertzuwachs. Basiszins 2026: 3,20 %. "
        "Bereits versteuerte Vorabpauschalen erhoehen die Anschaffungskosten und mindern spaeter den zu versteuernden "
        "Veraeusserungsgewinn - der Effekt ist also primaer ein Liquiditaets- und Zinseszinsnachteil."), "18 InvStG"),
    ("Teilfreistellung", (
        "Bei Aktienfonds mit einer Aktienquote ueber 50 % bleiben 30 % der Ertraege im Privatvermoegen steuerfrei. "
        "Effektive Belastung damit rund 18,5 % statt 26,375 %."), "20 InvStG"),
    ("Abgeltungsteuer", (
        "25 % zuzueglich Solidaritaetszuschlag und gegebenenfalls Kirchensteuer. Effektiv 26,375 % ohne bzw. rund "
        "27,8 bis 28,0 % mit Kirchensteuer. Anders als bei der nachgelagerten Besteuerung ist der Satz vom "
        "persoenlichen Grenzsteuersatz unabhaengig - der grosse Vorteil des freien Depots bei hohen Einkommen."),
     "32d, 43a EStG"),
    ("Sparer-Pauschbetrag", (
        "1.000 EUR je Person und Jahr steuerfreier Kapitalertrag, bei Zusammenveranlagung 2.000 EUR."), "20 Abs. 9 EStG"),

    ("H", "Auswertung", None, None),
    ("Barwert", (
        "Auf heute abgezinster Betrag, hier mit der Inflationsrate diskontiert. Damit sind alle Zahlen in Kaufkraft "
        "von heute vergleichbar - unerlaesslich, wenn Zahlungen 30 bis 50 Jahre auseinanderliegen."), None),
    ("Netto-Ertrag", (
        "Barwert aller Netto-Auszahlungen minus Barwert des gesamten Nettoaufwands. Die Hauptkennzahl: was eine Option "
        "ueber den eigenen Einsatz hinaus abwirft."), "Zeile 15 im Blatt 'Vergleich'"),
    ("Interner Zinsfuss (IRR)", (
        "Der Zinssatz, bei dem der Kapitalwert aller Zahlungsstroeme null wird, also die Nullstelle des Polynoms "
        "Summe CF_t / (1+i)^t. Direkt mit der angenommenen Bruttorendite des ETF vergleichbar: liegt der IRR darunter, "
        "frisst die Foerderarchitektur mehr als sie bringt."), "Blatt 'Cashflow'"),
    ("Annuitaetenfaktor", (
        "Faktor, der ein Kapital in eine konstante nominale Jahresentnahme ueber n Jahre umrechnet: "
        "a = r / (1 - (1+r)^-n). Wird fuer die Auszahlungsplaene von Altersvorsorgedepot und Privatdepot benutzt."),
     "Modellkonvention"),
    ("Sleeve", (
        "Hilfsbegriff des Modells: der Teil des Nettoaufwands, der den jaehrlichen Hoechstbetrag des "
        "Altersvorsorgevertrags von 6.840 EUR uebersteigt. Er wird in ein normales Depot geleitet und wie Option 3 "
        "besteuert, damit alle drei Optionen exakt denselben Nettoaufwand haben."), "Modellkonvention"),
    ("Vergleichshorizont", (
        "Alter, bis zu dem gerechnet wird (Default 85). Auszahlungsplaene enden hier. Eine lebenslange bAV-Rente zahlt "
        "darueber hinaus; dieser Betrag steht als Memo-Zeile im Blatt 'Vergleich' und geht NICHT in den Vergleich ein."),
     "Eingabe"),
]

r = 4
for item in GLOSS:
    if item[0] == "H":
        bw.cell(r, 2, item[1]).font = H2
        for c in (2, 3, 4):
            bw.cell(r, c).fill = GREYFILL
        r += 1
        continue
    term, expl, ref = item
    tc = bw.cell(r, 2, term)
    tc.font = BOLD
    tc.alignment = Alignment(vertical="top")
    tc.border = BOX
    ec = bw.cell(r, 3, expl)
    ec.font = Font(name=FONT, size=10)
    ec.alignment = Alignment(wrap_text=True, vertical="top")
    ec.border = BOX
    rc = bw.cell(r, 4, ref if ref else "")
    rc.font = Font(name=FONT, size=9, italic=True, color="595959")
    rc.alignment = Alignment(wrap_text=True, vertical="top")
    rc.border = BOX
    lines = expl.count("\n") + 1 + len(expl) // 100
    bw.row_dimensions[r].height = max(28, 13 * lines)
    r += 1

wb.save(OUT)
print("saved", OUT)
